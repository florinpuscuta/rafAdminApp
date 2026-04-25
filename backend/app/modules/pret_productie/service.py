"""Pret productie — parsing + upsert + list pentru tabelul production_prices.

Fisierul Excel asteptat are coloane: Cod produs | Denumire produs | Grupa |
Pret vanzare | Pret productie | Marja. Citim doar `Cod produs` + `Pret
productie`. Restul coloanelor sunt informationale (le ignoram). Match-ul cu
produsele canonice se face strict pe `Product.code` in tenant-ul curent.

Scope-ul ('adp' / 'sika') izoleaza cele doua liste — un upload SIKA NU
modifica preturile ADP si invers.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.pret_productie.models import ProductionPrice, ProductionPriceMonthly
from app.modules.products.models import Product
from app.modules.product_categories.models import ProductCategory


SCOPES = ("adp", "sika")


# --- Parser ----------------------------------------------------------------

@dataclass
class _ParsedRow:
    code: str
    price: Decimal


@dataclass
class _ParseOutcome:
    rows: list[_ParsedRow]
    invalid: int  # randuri cu cod gol / pret neparsabil


def _norm_header(s: object) -> str:
    return (str(s) if s is not None else "").strip().lower()


_CODE_HEADERS = {"cod produs", "cod", "code", "product code"}
_PRICE_HEADERS = {
    "pret productie", "pret productie (ron)", "pret_productie",
    "production price", "cost", "pret cost",
}


def parse_xlsx(content: bytes) -> _ParseOutcome:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    code_col: int | None = None
    price_col: int | None = None
    for raw_row in rows_iter:
        norm = [_norm_header(c) for c in raw_row]
        for i, h in enumerate(norm):
            if code_col is None and h in _CODE_HEADERS:
                code_col = i
            if price_col is None and h in _PRICE_HEADERS:
                price_col = i
        if code_col is not None and price_col is not None:
            break

    if code_col is None or price_col is None:
        raise ValueError(
            "Header-ul fisierului nu contine coloanele 'Cod produs' si "
            "'Pret productie'. Asigura-te ca primul rand cu antete contine "
            "aceste denumiri."
        )

    out: list[_ParsedRow] = []
    invalid = 0
    for raw_row in rows_iter:
        if raw_row is None:
            continue
        code_val = raw_row[code_col] if code_col < len(raw_row) else None
        price_val = raw_row[price_col] if price_col < len(raw_row) else None
        code = str(code_val).strip() if code_val is not None else ""
        if not code:
            continue
        try:
            if price_val is None or price_val == "":
                invalid += 1
                continue
            price = Decimal(str(price_val))
        except (InvalidOperation, ValueError):
            invalid += 1
            continue
        if price <= 0:
            invalid += 1
            continue
        out.append(_ParsedRow(code=code, price=price))

    return _ParseOutcome(rows=out, invalid=invalid)


# --- Upsert ----------------------------------------------------------------

async def upsert_prices(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    scope: str,
    parsed_rows: list[_ParsedRow],
    filename: str,
) -> tuple[int, int, list[str]]:
    """Replace total: sterge TOT scope-ul, apoi inserta produsele matchate.

    Returneaza (inserted, deleted_before, unmatched_codes). Al doilea camp
    e "deleted_before_insert" — randuri vechi sterse inainte de inserare,
    pentru parita semnatura cu varianta veche (inserted, updated).
    """
    if not parsed_rows:
        return 0, 0, []

    codes = [r.code for r in parsed_rows]
    res = await session.execute(
        select(Product.id, Product.code).where(
            Product.tenant_id == tenant_id, Product.code.in_(codes),
        )
    )
    code_to_pid: dict[str, UUID] = {row.code: row.id for row in res}

    matched_pairs: list[tuple[UUID, Decimal]] = []
    unmatched: list[str] = []
    for r in parsed_rows:
        pid = code_to_pid.get(r.code)
        if pid is None:
            unmatched.append(r.code)
        else:
            matched_pairs.append((pid, r.price))

    # Sterge tot scope-ul inainte sa inseram noul set — upload-ul e "replace"
    # total. Aceasta evita stale rows din upload-uri anterioare cu liste
    # diferite (ex: ai dat un fisier cu 1460 produse, apoi unul cu 235 — vrem
    # sa ramana 235, nu 1460).
    del_res = await session.execute(
        delete(ProductionPrice).where(
            ProductionPrice.tenant_id == tenant_id,
            ProductionPrice.scope == scope,
        )
    )
    deleted_before = del_res.rowcount or 0

    if not matched_pairs:
        await session.commit()
        return 0, deleted_before, unmatched

    now = datetime.utcnow()
    payload = [
        {
            "tenant_id": tenant_id,
            "scope": scope,
            "product_id": pid,
            "price": price,
            "last_imported_at": now,
            "last_imported_filename": filename,
        }
        for pid, price in matched_pairs
    ]
    await session.execute(pg_insert(ProductionPrice).values(payload))
    await session.commit()
    return len(matched_pairs), deleted_before, unmatched


# --- Read ------------------------------------------------------------------

async def get_summary(
    session: AsyncSession, tenant_id: UUID,
) -> dict[str, dict]:
    """Returneaza, per scope, count + ultima incarcare."""
    res = await session.execute(
        select(
            ProductionPrice.scope,
            func.count(ProductionPrice.id).label("c"),
            func.max(ProductionPrice.last_imported_at).label("max_ts"),
        )
        .where(ProductionPrice.tenant_id == tenant_id)
        .group_by(ProductionPrice.scope)
    )
    summary: dict[str, dict] = {
        s: {"count": 0, "last_imported_at": None, "last_imported_filename": None}
        for s in SCOPES
    }
    for row in res:
        summary[row.scope] = {
            "count": int(row.c),
            "last_imported_at": row.max_ts,
            "last_imported_filename": None,
        }

    # Pentru fiecare scope cu date, luam si filename-ul ultimei incarcari.
    for scope, info in summary.items():
        if info["last_imported_at"] is None:
            continue
        fr = await session.execute(
            select(ProductionPrice.last_imported_filename)
            .where(
                ProductionPrice.tenant_id == tenant_id,
                ProductionPrice.scope == scope,
                ProductionPrice.last_imported_at == info["last_imported_at"],
            )
            .limit(1)
        )
        info["last_imported_filename"] = fr.scalar_one_or_none()
    return summary


async def list_prices(
    session: AsyncSession, tenant_id: UUID, scope: str,
) -> list[dict]:
    res = await session.execute(
        select(
            ProductionPrice.product_id,
            ProductionPrice.price,
            Product.code,
            Product.name,
            ProductCategory.label,
        )
        .join(Product, Product.id == ProductionPrice.product_id)
        .outerjoin(ProductCategory, ProductCategory.id == Product.category_id)
        .where(
            ProductionPrice.tenant_id == tenant_id,
            ProductionPrice.scope == scope,
        )
        .order_by(Product.name)
    )
    return [
        {
            "product_id": row.product_id,
            "product_code": row.code,
            "product_name": row.name,
            "category_label": row.label,
            "price": row.price,
        }
        for row in res
    ]


async def reset_scope(
    session: AsyncSession, tenant_id: UUID, scope: str,
) -> int:
    res = await session.execute(
        delete(ProductionPrice).where(
            ProductionPrice.tenant_id == tenant_id,
            ProductionPrice.scope == scope,
        )
    )
    await session.commit()
    return res.rowcount or 0


# ─── Snapshot lunar ──────────────────────────────────────────────────

async def upsert_prices_monthly(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    scope: str,
    year: int,
    month: int,
    parsed_rows: list[_ParsedRow],
    filename: str,
) -> tuple[int, int, list[str]]:
    """Replace total al snapshot-ului pe (tenant, scope, year, month).
    Returneaza (inserted, deleted_before, unmatched_codes).
    """
    if not parsed_rows:
        return 0, 0, []
    codes = [r.code for r in parsed_rows]
    res = await session.execute(
        select(Product.id, Product.code).where(
            Product.tenant_id == tenant_id, Product.code.in_(codes),
        )
    )
    code_to_pid: dict[str, UUID] = {row.code: row.id for row in res}

    matched_pairs: list[tuple[UUID, Decimal]] = []
    unmatched: list[str] = []
    for r in parsed_rows:
        pid = code_to_pid.get(r.code)
        if pid is None:
            unmatched.append(r.code)
        else:
            matched_pairs.append((pid, r.price))

    # Replace total al snapshot-ului lunar pentru (tenant, scope, year, month).
    del_res = await session.execute(
        delete(ProductionPriceMonthly).where(
            ProductionPriceMonthly.tenant_id == tenant_id,
            ProductionPriceMonthly.scope == scope,
            ProductionPriceMonthly.year == year,
            ProductionPriceMonthly.month == month,
        )
    )
    deleted_before = del_res.rowcount or 0

    if not matched_pairs:
        await session.commit()
        return 0, deleted_before, unmatched

    now = datetime.utcnow()
    payload = [
        {
            "tenant_id": tenant_id,
            "scope": scope,
            "product_id": pid,
            "year": year,
            "month": month,
            "price": price,
            "last_imported_at": now,
            "last_imported_filename": filename,
        }
        for pid, price in matched_pairs
    ]
    await session.execute(pg_insert(ProductionPriceMonthly).values(payload))
    await session.commit()
    return len(matched_pairs), deleted_before, unmatched


async def get_monthly_summary(
    session: AsyncSession, tenant_id: UUID,
) -> dict[str, list[dict]]:
    """Per scope, lista lunilor cu snapshot incarcat (year, month, count, ts)."""
    res = await session.execute(
        select(
            ProductionPriceMonthly.scope,
            ProductionPriceMonthly.year,
            ProductionPriceMonthly.month,
            func.count(ProductionPriceMonthly.id).label("c"),
            func.max(ProductionPriceMonthly.last_imported_at).label("max_ts"),
        )
        .where(ProductionPriceMonthly.tenant_id == tenant_id)
        .group_by(
            ProductionPriceMonthly.scope,
            ProductionPriceMonthly.year,
            ProductionPriceMonthly.month,
        )
        .order_by(
            ProductionPriceMonthly.scope,
            ProductionPriceMonthly.year.desc(),
            ProductionPriceMonthly.month.desc(),
        )
    )
    out: dict[str, list[dict]] = {s: [] for s in SCOPES}
    for row in res:
        out.setdefault(row.scope, []).append({
            "year": int(row.year),
            "month": int(row.month),
            "count": int(row.c),
            "last_imported_at": row.max_ts,
        })
    return out


async def list_prices_monthly(
    session: AsyncSession,
    tenant_id: UUID,
    scope: str,
    year: int,
    month: int,
) -> list[dict]:
    res = await session.execute(
        select(
            ProductionPriceMonthly.product_id,
            ProductionPriceMonthly.price,
            Product.code,
            Product.name,
            ProductCategory.label,
        )
        .join(Product, Product.id == ProductionPriceMonthly.product_id)
        .outerjoin(ProductCategory, ProductCategory.id == Product.category_id)
        .where(
            ProductionPriceMonthly.tenant_id == tenant_id,
            ProductionPriceMonthly.scope == scope,
            ProductionPriceMonthly.year == year,
            ProductionPriceMonthly.month == month,
        )
        .order_by(Product.name)
    )
    return [
        {
            "product_id": row.product_id,
            "product_code": row.code,
            "product_name": row.name,
            "category_label": row.label,
            "price": row.price,
        }
        for row in res
    ]


async def reset_scope_monthly(
    session: AsyncSession,
    tenant_id: UUID,
    scope: str,
    year: int,
    month: int,
) -> int:
    res = await session.execute(
        delete(ProductionPriceMonthly).where(
            ProductionPriceMonthly.tenant_id == tenant_id,
            ProductionPriceMonthly.scope == scope,
            ProductionPriceMonthly.year == year,
            ProductionPriceMonthly.month == month,
        )
    )
    await session.commit()
    return res.rowcount or 0
