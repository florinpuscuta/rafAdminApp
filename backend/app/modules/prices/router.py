import csv
import io
from datetime import datetime, timezone
from decimal import Decimal
from uuid import UUID

from fastapi import Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.api import APIRouter
from app.core.db import get_session
from app.modules.auth.deps import get_current_tenant_id
from app.modules.prices import service as prices_service
from app.modules.prices.schemas import (
    CrossKaPrice,
    CrossKaResponse,
    CrossKaRow,
    KaRetailResponse,
    KaRetailRow,
    PriceComparisonResponse,
    PriceComparisonRow,
    PriceComparisonSummary,
    Pret3NetClient,
    Pret3NetProduct,
    Pret3NetResponse,
    PropunereRow,
    PropuneriResponse,
)


def _parse_months(months: str | None) -> list[int] | None:
    """Parsez param `months=1,2,3` → [1,2,3]. Ignor valori invalide."""
    if not months:
        return None
    out: list[int] = []
    for part in months.split(","):
        p = part.strip()
        if not p:
            continue
        try:
            m = int(p)
            if 1 <= m <= 12:
                out.append(m)
        except ValueError:
            continue
    return out or None

router = APIRouter(prefix="/api/prices", tags=["prices"])


@router.get("/ka-vs-tt", response_model=PriceComparisonResponse)
async def ka_vs_tt(
    year: int | None = Query(None),
    month: int | None = Query(None),
    category: str | None = Query(None),
    product_id: UUID | None = Query(None, alias="productId"),
    min_qty: Decimal = Query(Decimal("0.01"), alias="minQty", ge=0),
    limit: int = Query(500, ge=1, le=2000),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """
    Compară prețul mediu Adeplast pe KA (Key Accounts) vs TT (Traditional Trade)
    pentru fiecare produs. Include doar produsele vândute în AMBELE canale cu
    cantitate ≥ `minQty`.

    Sursă: `raw_sales` — nu există tabel dedicat, calculul se face la cerere.
    """
    rows = await prices_service.compare_ka_vs_tt(
        session, tenant_id,
        year=year, month=month, category=category,
        product_id=product_id, min_qty=min_qty, limit=limit,
    )
    summary = await prices_service.summary_ka_vs_tt(
        session, tenant_id, year=year, month=month, category=category,
    )
    return PriceComparisonResponse(
        summary=PriceComparisonSummary(**summary),
        rows=[PriceComparisonRow(**r) for r in rows],
    )


@router.get("/ka-vs-tt/export")
async def ka_vs_tt_export(
    year: int | None = Query(None),
    month: int | None = Query(None),
    category: str | None = Query(None),
    product_id: UUID | None = Query(None, alias="productId"),
    min_qty: Decimal = Query(Decimal("0.01"), alias="minQty", ge=0),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """CSV export cu toate rândurile comparației — pentru analiza offline."""
    rows = await prices_service.compare_ka_vs_tt(
        session, tenant_id,
        year=year, month=month, category=category,
        product_id=product_id, min_qty=min_qty, limit=2000,
    )

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "description", "product_code", "category",
        "ka_price", "ka_qty", "ka_sales",
        "tt_price", "tt_qty", "tt_sales",
        "delta_abs", "delta_pct",
    ])
    for r in rows:
        w.writerow([
            r["description"] or "",
            r["product_code"] or "",
            r["category"] or "",
            f"{r['ka_price']:.4f}" if r["ka_price"] is not None else "",
            f"{r['ka_qty']:.3f}",
            f"{r['ka_sales']:.2f}",
            f"{r['tt_price']:.4f}" if r["tt_price"] is not None else "",
            f"{r['tt_qty']:.3f}",
            f"{r['tt_sales']:.2f}",
            f"{r['delta_abs']:.4f}" if r["delta_abs"] is not None else "",
            f"{r['delta_pct']:.2f}" if r["delta_pct"] is not None else "",
        ])
    buf.seek(0)

    filename = f"ka-vs-tt-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ────────────────────────────────────────────────────────────────────────────
# /prices/own — Prețuri cross-KA (brand propriu)
# ────────────────────────────────────────────────────────────────────────────


@router.get("/own", response_model=CrossKaResponse)
async def cross_ka_own(
    year: int | None = Query(None),
    months: str | None = Query(None, description="CSV, ex: 1,2,3"),
    category: str | None = Query(None),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Pentru fiecare produs vândut la cel puțin 2 KA, afișează prețul mediu
    per rețea + min/max/spread. Util pentru a spoti dispersia de preț."""
    data = await prices_service.cross_ka_own(
        session, tenant_id, year=year, months=_parse_months(months), category=category,
    )
    rows = [
        CrossKaRow(
            description=r["description"],
            product_code=r["product_code"],
            category=r["category"],
            prices={k: CrossKaPrice(**v) for k, v in r["prices"].items()},
            min_price=r["min_price"],
            max_price=r["max_price"],
            spread_pct=r["spread_pct"],
            n_stores=r["n_stores"],
        )
        for r in data["rows"]
    ]
    return CrossKaResponse(ka_clients=data["ka_clients"], rows=rows)


# ────────────────────────────────────────────────────────────────────────────
# /prices/pret3net — Preț 3 Net Comp KA
# ────────────────────────────────────────────────────────────────────────────


@router.get("/pret3net", response_model=Pret3NetResponse)
async def pret3net(
    year: int | None = Query(None),
    months: str | None = Query(None, description="CSV, ex: 1,2,3"),
    company: str = Query("adeplast", description="adeplast|sika|sikadp"),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Preț mediu (sales/qty) per produs, per client KA, grupat pe categorie.
    `company` filtrează: adeplast exclude Sika, sika include doar Sika.
    """
    data = await prices_service.pret3net(
        session, tenant_id, year=year, months=_parse_months(months), company=company,
    )
    categories = {
        cat: [
            Pret3NetProduct(
                description=p["description"],
                clients={k: Pret3NetClient(**v) for k, v in p["clients"].items()},
                total_sales=p["total_sales"],
                total_qty=p["total_qty"],
            )
            for p in prods
        ]
        for cat, prods in data["categories"].items()
    }
    return Pret3NetResponse(year=data["year"], ka_clients=data["ka_clients"], categories=categories)


# ────────────────────────────────────────────────────────────────────────────
# /prices/propuneri — Propuneri Listare KA
# ────────────────────────────────────────────────────────────────────────────


@router.get("/propuneri", response_model=PropuneriResponse)
async def propuneri_listare(
    year: int | None = Query(None),
    months: str | None = Query(None, description="CSV, ex: 1,2,3"),
    company: str = Query("adeplast", description="adeplast|sika|sikadp"),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Pentru fiecare KA, produsele vândute la alte KA dar NU la acesta,
    împreună cu prețul minim dintre celelalte rețele (ca referință).

    `company` filtrează: adeplast exclude Sika, sika include doar Sika.
    Produsele de marca privată sunt excluse ÎNTOTDEAUNA (nu se propun spre
    listare la alte rețele — nu are sens comercial).
    """
    data = await prices_service.propuneri_listare(
        session, tenant_id, year=year, months=_parse_months(months), company=company,
    )
    suggestions = {
        ka: [PropunereRow(**item) for item in items]
        for ka, items in data["suggestions"].items()
    }
    return PropuneriResponse(
        year=data["year"], ka_clients=data["ka_clients"], suggestions=suggestions,
    )


# ────────────────────────────────────────────────────────────────────────────
# /prices/ka-retail — Top produse KA vs Retail
# ────────────────────────────────────────────────────────────────────────────


@router.get("/ka-retail", response_model=KaRetailResponse)
async def ka_vs_retail(
    year: int | None = Query(None),
    months: str | None = Query(None, description="CSV, ex: 1,2,3"),
    category: str | None = Query(None),
    limit: int = Query(15, ge=1, le=200),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Top N produse vândute și pe KA și pe Retail (channel strict 'RETAIL'),
    cu prețuri medii și diferență procentuală. Distinct de /ka-vs-tt, care
    tratează orice non-KA ca TT."""
    rows = await prices_service.ka_vs_retail(
        session, tenant_id,
        year=year, months=_parse_months(months), category=category, limit=limit,
    )
    return KaRetailResponse(rows=[KaRetailRow(**r) for r in rows])


# ─────────────────────────────────────────────────────────────────────────
# Port LEGACY Prețuri Comparative (grid 4-rețele) + Adeplast/Sika cross-KA
# Endpoint-uri adaptate din `adeplast-dashboard/routes/pricing.py`.
# Datele vin din tabelele `price_grid` + `price_grid_meta` (import one-off
# din legacy `adeplast_ka.db` + `sika_ka.db`).
# ─────────────────────────────────────────────────────────────────────────

_VALID_COMPANIES = {"adeplast", "sika"}
_LEGACY_KA_STORES = ["Dedeman", "Leroy", "Hornbach", "Brico"]


@router.get("/grid/{store}")
async def api_legacy_price_grid(
    store: str,
    company: str = Query("adeplast"),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Port 1:1 al legacy GET /api/price_grid/<store>.

    Returnează grid-ul Prețuri Comparative pentru o rețea (Dedeman/Leroy/
    Hornbach/Brico) + meta (date_prices, branduri).
    """
    company = company.lower()
    if company not in _VALID_COMPANIES:
        company = "adeplast"
    data = await prices_service.get_price_grid(
        session, tenant_id, store=store, company=company,
    )
    return {"ok": True, **data}


@router.get("/grid-export.xlsx")
async def api_legacy_price_grid_export(
    company: str = Query("adeplast"),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Export Excel al tuturor grid-urilor (Dedeman/Leroy/Hornbach/Brico) —
    câte un sheet per magazin. Fiecare rând produs = 1 linie; fiecare brand
    are 2 coloane (Produs + Preț). Formatare: header bold, preț cu 2 zecimale.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from fastapi.responses import Response as FResponse

    company = company.lower()
    if company not in _VALID_COMPANIES:
        company = "adeplast"

    stores = ["Dedeman", "Leroy", "Hornbach", "Brico"]
    wb = Workbook()
    wb.remove(wb.active)

    anchor_brand = "ADEPLAST" if company == "adeplast" else "Sika"

    # Fill-uri pe POZIȚIONARE preț (vs ancora):
    # VERDE  = ancora e mai ieftin → BINE pozitionat (competitorul e scump)
    # ROȘU   = ancora e mai scump  → SLAB pozitionat (competitorul e ieftin)
    # NEUTRU = diferență ≤ 5%
    FILL_GREEN = PatternFill("solid", fgColor="D1FAE5")   # green-100
    FILL_RED = PatternFill("solid", fgColor="FEE2E2")     # red-100
    FILL_NEUTRAL = PatternFill("solid", fgColor="F8FAFC") # slate-50
    FILL_ANCHOR = PatternFill("solid", fgColor="DBEAFE")  # blue-100 — ancora highlight
    FONT_GREEN = Font(color="166534", bold=True, size=11)
    FONT_RED = Font(color="991B1B", bold=True, size=11)
    FONT_NEUTRAL = Font(color="334155", size=11)
    FONT_ANCHOR = Font(color="1E40AF", bold=True, size=11)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    header_fill_anchor = PatternFill("solid", fgColor="1E40AF")  # blue-800 pentru ancoră
    sub_fill = PatternFill("solid", fgColor="F1F5F9")
    sub_font = Font(bold=True, size=10, color="334155")

    def positioning_for(anchor_price, comp_price, tolerance=0.05):
        """Returnează (fill, font) pe baza comparației preț vs ancoră."""
        if comp_price is None or anchor_price is None or anchor_price <= 0:
            return FILL_NEUTRAL, FONT_NEUTRAL
        ratio = comp_price / anchor_price
        if ratio > 1 + tolerance:
            # Competitor e mai scump → ancora e bine pozitionată
            return FILL_GREEN, FONT_GREEN
        if ratio < 1 - tolerance:
            # Competitor e mai ieftin → ancora e slab pozitionată
            return FILL_RED, FONT_RED
        return FILL_NEUTRAL, FONT_NEUTRAL

    for store in stores:
        data = await prices_service.get_price_grid(
            session, tenant_id, store=store, company=company,
        )
        meta = data.get("meta") or {}
        rows = data.get("rows") or []
        brands: list[str] = meta.get("brands") or []

        ws = wb.create_sheet(store[:31])  # Excel sheet name ≤ 31 chars

        # Header 1: "#" + brand (col-span 2)
        ws.cell(row=1, column=1, value="#").font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value="Grupă").font = header_font_white
        ws.cell(row=1, column=2).fill = header_fill
        col = 3
        for b in brands:
            cell = ws.cell(row=1, column=col, value=b)
            cell.font = header_font_white
            # Ancora = albastru închis, restul = slate-800
            cell.fill = header_fill_anchor if b == anchor_brand or b.upper() == anchor_brand.upper() else header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
            col += 2

        # Header 2: sub (Produs / Preț)
        ws.cell(row=2, column=1).fill = sub_fill
        ws.cell(row=2, column=2).fill = sub_fill
        col = 3
        for _ in brands:
            c1 = ws.cell(row=2, column=col, value="Produs")
            c1.font = sub_font; c1.fill = sub_fill
            c2 = ws.cell(row=2, column=col + 1, value="Preț (RON)")
            c2.font = sub_font; c2.fill = sub_fill
            c2.alignment = Alignment(horizontal="right")
            col += 2

        # Date — culoare pe POZIȚIONARE (verde/roșu/neutru)
        for i, r in enumerate(rows, start=3):
            brand_data = r.get("brand_data") or {}
            ws.cell(row=i, column=1, value=r.get("row_num") or (r.get("row_idx", 0) + 1))
            ws.cell(row=i, column=2, value=r.get("group_label") or "")

            # Extrage prețul ancorei pentru rândul curent
            anchor_obj = brand_data.get(anchor_brand) or brand_data.get(anchor_brand.upper()) or {}
            anchor_price = None
            if isinstance(anchor_obj, dict):
                try:
                    anchor_price = float(anchor_obj.get("pret")) if anchor_obj.get("pret") else None
                except (TypeError, ValueError):
                    anchor_price = None

            col = 3
            for b in brands:
                cell_data = brand_data.get(b) or brand_data.get(b.upper()) or {}
                prod = cell_data.get("prod") if isinstance(cell_data, dict) else None
                pret = cell_data.get("pret") if isinstance(cell_data, dict) else None
                try:
                    pret_val = float(pret) if pret is not None else None
                except (TypeError, ValueError):
                    pret_val = None

                # Ancora: fill albastru deschis, font bold albastru
                if b == anchor_brand or b.upper() == anchor_brand.upper():
                    cell_fill = FILL_ANCHOR
                    cell_font = FONT_ANCHOR
                else:
                    # Competitor: culoare funcție de comparație cu ancoră
                    cell_fill, cell_font = positioning_for(anchor_price, pret_val)

                prod_cell = ws.cell(row=i, column=col, value=prod or "")
                prod_cell.fill = cell_fill
                prod_cell.font = Font(size=10, color=cell_font.color.rgb if cell_font.color else "334155")

                pret_cell = ws.cell(row=i, column=col + 1)
                pret_cell.fill = cell_fill
                pret_cell.font = cell_font
                pret_cell.alignment = Alignment(horizontal="right")
                if pret_val is not None:
                    pret_cell.value = pret_val
                    pret_cell.number_format = '#,##0.00'
                col += 2

        # Lățimi coloane — brand name = header cell centered
        ws.column_dimensions["A"].width = 5   # #
        ws.column_dimensions["B"].width = 14  # Grupă
        for idx in range(len(brands)):
            prod_col = chr(ord("A") + 2 + idx * 2)  # safe <= G (7+1)
            pret_col_idx = 2 + 1 + idx * 2 + 1  # 1-based
            # Use openpyxl's get_column_letter for safety on >26
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(3 + idx * 2)].width = 34
            ws.column_dimensions[get_column_letter(3 + idx * 2 + 1)].width = 12

        ws.freeze_panes = "C3"

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    company_label = "adeplast" if company == "adeplast" else "sika"
    filename = f"preturi-comparative-{company_label}.xlsx"
    return FResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.put("/grid/{store}/cell")
async def api_legacy_price_grid_cell(
    store: str,
    body: dict,
    company: str = Query("adeplast"),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Port 1:1 al legacy PUT /api/price_grid/<store>/cell.

    Update manual al unei celule. Body: { row_idx, brand, prod?, pret? }.
    Când `pret` e actualizat → setează ai_status='manual' (punct albastru).
    """
    company = company.lower()
    if company not in _VALID_COMPANIES:
        company = "adeplast"
    try:
        row_idx = int(body.get("row_idx"))
        brand = (body.get("brand") or "").strip()
    except (ValueError, TypeError):
        from fastapi import HTTPException, status
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="row_idx/brand invalid")
    if not brand:
        from fastapi import HTTPException, status
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="brand lipseste")

    updated = await prices_service.update_price_grid_cell(
        session, tenant_id, store=store, company=company,
        row_idx=row_idx, brand=brand,
        prod=body.get("prod"), pret=body.get("pret"),
    )
    if not updated:
        from fastapi import HTTPException, status
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="row not found")
    return {"ok": True}


# AI update endpoints — port 1:1 din legacy routes/pricing.py:645-701
@router.post("/grid/{store}/ai_update/start")
async def api_ai_update_start(
    store: str,
    body: dict | None = None,
    company: str = Query("adeplast"),
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """Port al legacy POST /api/price_grid/<store>/ai_update/start.

    Body (optional): { "provider": "grok"|"anthropic"|"openai" }
    Cheile se citesc din env: XAI_API_KEY, ANTHROPIC_API_KEY, OPENAI_API_KEY.
    """
    from app.modules.prices import ai_update_service as ai_svc
    from fastapi import HTTPException, status as http_status
    company = company.lower()
    if company not in _VALID_COMPANIES:
        company = "adeplast"
    preferred = (body or {}).get("provider") if body else None
    try:
        job_id, total, provider = await ai_svc.start_update_job(
            tenant_id, store=store, company=company,
            preferred_provider=preferred,
        )
        return {"ok": True, "job_id": job_id, "total": total,
                "company": company, "provider": provider}
    except RuntimeError as e:
        # active job or missing key
        active = await ai_svc.get_active_job(tenant_id, store, company)
        raise HTTPException(
            http_status.HTTP_409_CONFLICT,
            detail={"ok": False, "error": str(e),
                    "job_id": active.get("job_id") if active else None},
        )


@router.get("/grid/{store}/ai_update/status")
async def api_ai_update_status(
    store: str,
    company: str = Query("adeplast"),
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """Status-ul job-ului curent/latest pentru rețea + companie."""
    from app.modules.prices import ai_update_service as ai_svc
    company = company.lower()
    if company not in _VALID_COMPANIES:
        company = "adeplast"
    active = await ai_svc.get_active_job(tenant_id, store, company)
    if active:
        return {"ok": True, "job": active, "is_active": True}
    latest = await ai_svc.get_latest_job(tenant_id, store, company)
    return {"ok": True, "job": latest, "is_active": False}


@router.get("/grid/ai_update/{job_id}")
async def api_ai_update_job(
    job_id: str,
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """Status-ul unui job specific (polling)."""
    from app.modules.prices import ai_update_service as ai_svc
    from fastapi import HTTPException, status as http_status
    job = await ai_svc.get_job_progress(job_id)
    if not job:
        raise HTTPException(http_status.HTTP_404_NOT_FOUND, detail="job not found")
    return {"ok": True, "job": job}


@router.post("/grid/ai_update/{job_id}/cancel")
async def api_ai_update_cancel(
    job_id: str,
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """Cancel un job în curs."""
    from app.modules.prices import ai_update_service as ai_svc
    await ai_svc.cancel_job(job_id)
    return {"ok": True}


@router.get("/own_cross_ka")
async def api_legacy_own_cross_ka(
    company: str = Query("adeplast"),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Port 1:1 al legacy GET /api/price_grid/own_cross_ka.

    Tabel pivot cu produsele proprii (Adeplast sau Sika) × 4 rețele KA.
    Pentru fiecare produs: min/max/spread% între magazine.
    """
    company = company.lower()
    if company not in _VALID_COMPANIES:
        company = "adeplast"
    data = await prices_service.get_own_cross_ka(
        session, tenant_id, company=company,
    )
    return {"ok": True, **data}
