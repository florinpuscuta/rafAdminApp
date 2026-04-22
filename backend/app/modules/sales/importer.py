"""
Parser Excel → RawSale pentru upload-ul ADP.

Acceptă aliasuri multiple (RO+EN) pentru fiecare coloană și face header
auto-detect pe primele 20 rânduri — fișierele reale adesea au rânduri de
titlu / merged cells / rânduri goale înainte de header.

Output: (rows_to_insert, errors).

NOTĂ arhitecturală: raw_sales păstrează string-urile brute (client, agent,
product_code/name, category_code). Rezolvarea către entități canonice
(Store/Agent/Product + category/brand prin FK-uri) se face după insert în
service.bulk_insert via resolve_map — importer-ul nu se atinge de canonic.
"""
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

# ── Aliasuri pe câmp logic ────────────────────────────────────────────────
# Ordine în liste = ordinea de preferință. Normalizare: lower + strip +
# spațiu/punct/dash → underscore, diacritice → ASCII.

_ALIASES: dict[str, tuple[str, ...]] = {
    "year": ("year", "an", "anul", "yr"),
    "month": ("month", "luna", "lun", "mth", "mo"),
    "client": ("client", "firma", "company", "customer", "cust"),
    "ship_to": (
        "ship_to", "ship_to_code", "shipto",
        "punct_de_lucru", "punct_lucru",
    ),
    "amount": (
        "amount", "sales", "vanzari", "valoare", "val",
        "net_sales", "total", "suma",
    ),
    "quantity": (
        "quantity", "cantitate", "cant", "qty", "mc",
        "billed_qty", "kg",
    ),
    "channel": ("channel", "canal", "target_market", "piata"),
    "product_code": (
        "product_code", "cod_articol", "cod_produs", "sku", "material",
    ),
    "product_name": (
        "product_name", "descriere", "description", "material_desc",
        "denumire", "nume_produs",
    ),
    "category_code": (
        "category_code", "product_category", "product_group",
        "cod_categorie_articol",
    ),
    "agent": (
        "agent", "responsabil", "sales_rep", "rep",
        "agent_vanzari", "reprezentant",
    ),
}

REQUIRED_FIELDS = {"year", "month", "client", "amount"}

# Invers-map: alias → câmp logic (pentru lookup rapid după normalizare).
_ALIAS_TO_FIELD: dict[str, str] = {
    alias: field for field, aliases in _ALIASES.items() for alias in aliases
}


_DIACRITICS = str.maketrans(
    "ăâîșşțţĂÂÎȘŞȚŢáéíóúÁÉÍÓÚ",
    "aaissttaaisstTaeiouaeiou",
)


def _norm_header(value: Any) -> str:
    """
    Normalizare agresivă: lower + strip + înlocuiește diacritice + convertește
    spații/dash/punct/slash la underscore. "Cod Grup 2 Articol" → "cod_grup_2_articol".
    """
    if value is None:
        return ""
    s = str(value).strip().lower().translate(_DIACRITICS)
    out = []
    for ch in s:
        if ch.isalnum():
            out.append(ch)
        else:
            out.append("_")
    # colapsez underscores multiple
    normalized = "_".join(filter(None, "".join(out).split("_")))
    return normalized


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _detect_header(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """
    Caută printre primele 20 rânduri un rând care conține cel puțin 3 din
    câmpurile obligatorii (year/month/client/amount). Returnează (row_idx,
    {field: col_idx}) sau None dacă nu găsește.
    """
    for row_idx, row in enumerate(rows[:20]):
        col_to_field: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            field = _ALIAS_TO_FIELD.get(_norm_header(cell))
            if field and field not in col_to_field.values():
                col_to_field[col_idx] = field
        fields_found = set(col_to_field.values())
        if len(fields_found & REQUIRED_FIELDS) >= 3:
            return row_idx, {field: col for col, field in col_to_field.items()}
    return None


# Nume de sheet-uri preferate pentru raw-data. Ordinea contează.
# Sheet-urile de tip "Alocare" / "Marca Privata" / "Sheet1" (pivot-uri) sunt
# ignorate — sunt de obicei rapoarte agregate, nu raw.
_PREFERRED_SHEETS = [
    "!raportvanzarilunatotal",
    "raportvanzari",
    "raport_vanzari",
    "vanzari_detaliate",
    "raw_sales",
    "vanzari",
    "data",
    "raport",
]

_IGNORED_SHEETS_EXACT = {
    "alocare",
    "mapare",
    "marca privata",
    "sheet1",
}


# ── Alocare sheet: Client | Ship-to | ClientShip | Agent ─────────────────
# Sursa de adevăr pentru canonicals: transformăm fiecare rând într-un
# Store + Agent + StoreAlias + AgentAlias + AgentStoreAssignment.

_ALOCARE_ALIASES: dict[str, tuple[str, ...]] = {
    "client": ("client", "firma", "company", "customer"),
    "ship_to": ("ship_to", "ship_to_code", "shipto", "punct_de_lucru"),
    "agent": ("agent", "responsabil", "sales_rep", "reprezentant"),
    # "clientship" (client+ship concatenat) — ignorat, îl sintetizăm noi.
}

_ALOCARE_REQUIRED = {"client", "ship_to", "agent"}

_ALOCARE_SHEET_NAMES = ("alocare", "mapare")


def _alocare_alias_to_field(alias: str) -> str | None:
    for field, aliases in _ALOCARE_ALIASES.items():
        if alias in aliases:
            return field
    return None


def parse_alocare_sheet(content: bytes) -> list[dict[str, str]]:
    """
    Parse sheet-ul de mapare (nume tipic: "Alocare") cu triplete
    (Client, Ship-to, Agent). Returnează lista de dicționare cu:
        {raw_client, raw_ship_to, combined_key, agent_name}

    Unde `combined_key = "{raw_client} | {raw_ship_to}"` — formatul folosit
    în raw_sales.client după sinteza din parse_xlsx, ca să se potrivească
    la rezolvarea via alias.

    Dacă nu găsim un sheet Alocare, întoarce listă goală (nu-i eroare).
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = None
    for sheet in wb.worksheets:
        name = sheet.title.lower().strip()
        if any(key in name for key in _ALOCARE_SHEET_NAMES):
            ws = sheet
            break
    if ws is None:
        return []

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    # Detectare header în primele 20 rânduri (doar pentru Alocare).
    header_row_idx: int | None = None
    col_index: dict[str, int] = {}
    for row_idx, row in enumerate(all_rows[:20]):
        partial: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            field = _alocare_alias_to_field(_norm_header(cell))
            if field and field not in partial.values():
                partial[field] = col_idx
        if _ALOCARE_REQUIRED <= set(partial.keys()):
            header_row_idx = row_idx
            col_index = {v_field: col for v_field, col in
                         zip(partial.keys(), partial.values())}
            # flip (field→col) din structura inversă:
            col_index = {field: col for field, col in partial.items()}
            break
    if header_row_idx is None:
        return []

    out: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for row in all_rows[header_row_idx + 1 :]:
        if row is None or all(v is None or v == "" for v in row):
            continue

        def get(field: str, _row=row) -> Any:
            idx = col_index.get(field)
            return None if idx is None else (_row[idx] if idx < len(_row) else None)

        raw_client = _as_str(get("client"))
        raw_ship_to = _as_str(get("ship_to"))
        agent_name = _as_str(get("agent"))
        if not raw_client or not raw_ship_to or not agent_name:
            continue
        key = (raw_client, raw_ship_to, agent_name)
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "raw_client": raw_client,
            "raw_ship_to": raw_ship_to,
            "combined_key": f"{raw_client} | {raw_ship_to}",
            "agent_name": agent_name,
        })
    return out


def _pick_sheet(wb) -> Any:
    """
    Selectează sheet-ul care conține datele brute. Preferăm nume cunoscute,
    altfel primul care trece header-detect. Sheet-urile de tip pivot sunt
    skipate explicit.
    """
    # Mapez nume → sheet obiect (lowercase).
    sheets_by_name = {s.title.lower().strip(): s for s in wb.worksheets}
    for preferred in _PREFERRED_SHEETS:
        if preferred in sheets_by_name:
            return sheets_by_name[preferred]

    # Fallback: primul sheet care nu-i ignorat și are header detectabil.
    for ws in wb.worksheets:
        if ws.title.lower().strip() in _IGNORED_SHEETS_EXACT:
            continue
        sample = list(ws.iter_rows(values_only=True, max_row=20))
        if _detect_header(sample) is not None:
            return ws

    return wb.active


def parse_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if not wb.sheetnames:
        return [], ["Fișier Excel gol (niciun sheet)"]

    ws = _pick_sheet(wb)
    if ws is None:
        return [], ["Fișier Excel gol"]

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], [f"Sheet-ul '{ws.title}' e gol"]

    detected = _detect_header(all_rows)
    if detected is None:
        return [], [
            f"Sheet '{ws.title}': nu s-a găsit un rând de header valid în "
            "primele 20 rânduri. Verifică dacă Excel-ul conține coloane "
            "pentru: An/Year, Lună/Month, Client/Firma, Vânzări/Amount."
        ]

    header_row_idx, col_index = detected
    missing = REQUIRED_FIELDS - set(col_index.keys())
    if missing:
        return [], [
            f"Sheet '{ws.title}': lipsesc coloanele obligatorii: "
            + ", ".join(sorted(missing))
        ]

    rows_out: list[dict[str, Any]] = []
    errors: list[str] = []
    has_ship_to = "ship_to" in col_index

    for abs_idx, row in enumerate(all_rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if row is None or all(v is None or v == "" for v in row):
            continue

        def get(field: str, _row=row) -> Any:
            idx = col_index.get(field)
            return None if idx is None else (_row[idx] if idx < len(_row) else None)

        year = _as_int(get("year"))
        month = _as_int(get("month"))
        raw_client = _as_str(get("client"))
        amount = _as_decimal(get("amount"))

        if year is None or month is None or raw_client is None or amount is None:
            errors.append(
                f"Linia {abs_idx}: year/month/client/amount invalide sau lipsă"
            )
            continue
        if not (1 <= month <= 12):
            errors.append(f"Linia {abs_idx}: luna {month} în afara intervalului 1-12")
            continue
        if not (2000 <= year <= 2100):
            errors.append(f"Linia {abs_idx}: anul {year} în afara intervalului 2000-2100")
            continue

        # Sinteza client + ship_to într-o singură cheie (pattern „ClientShip"
        # din Alocare-ul legacy). Store-ul canonic e la nivel de punct de
        # livrare, nu de firma-mamă — așa rezolvă alias-urile unic.
        client_key = raw_client
        if has_ship_to:
            ship_to = _as_str(get("ship_to"))
            if ship_to:
                client_key = f"{raw_client} | {ship_to}"

        # Dacă Excel-ul nu are un product_code explicit dar are Description,
        # folosim Description ca product_code (pentru alias resolver).
        raw_code = _as_str(get("product_code"))
        raw_name = _as_str(get("product_name"))
        if raw_code is None and raw_name is not None:
            raw_code = raw_name

        rows_out.append(
            {
                "year": year,
                "month": month,
                "client": client_key,
                "channel": _as_str(get("channel")),
                "product_code": raw_code,
                "product_name": raw_name,
                "category_code": _as_str(get("category_code")),
                "amount": amount,
                "quantity": _as_decimal(get("quantity")),
                "agent": _as_str(get("agent")),
            }
        )

    return rows_out, errors
