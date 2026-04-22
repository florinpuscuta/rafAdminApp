"""
Parser Excel → RawSale pentru exportul Sika (DIY).

Forma fișierului e fundamental diferită de ADP:
  - câte un sheet per lună (nume "IAN 26", " APR 25", "DECEMBRIE 25"…);
  - în loc de coloane An/Lună, luna + anul sunt codificate în HEADER-ul
    coloanelor de Net Sales ("JAN 2025 …", "JAN 2026 …");
  - fiecare rând aduce DOUĂ valori — anul Y1 (an-1) și anul Y2 (anul curent
    al sheet-ului) — generăm 2 rânduri RawSale când ambele sunt populate.
  - nu există coloană "Agent" în sursă; agentul se rezolvă 100% din Raf
    mapping la backfill (primar pe cod ship-to, fallback pe nume).

Layout coloane (0-indexat):
   0  Customer (cod numeric)
   1  Customer (nume)               → client
   2  Ship-to Party (nume)          → ship_to
   3  Ship-to Party (cod numeric)   → client_code (match primar backfill)
   4  > Target Market
   5  Material (cod)                → product_code
   6  Material (denumire)           → product_name
   7  > Local Pr. Hier. 1           → category_code
   8  > Local Pr. Hier. 2 (brand)
   9  Net Sales year-1              → amount (emite rând cu year=Y1)
  10  Net Sales year                → amount (emite rând cu year=Y2)
  11  Billed Qty year-1             → quantity Y1
  12  Billed Qty year               → quantity Y2
"""
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook


_MONTH_TOKENS: dict[str, int] = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

_NET_SALES_HEADER_RE = re.compile(
    r"\b(?P<mon>JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+(?P<year>\d{4})\b",
    re.IGNORECASE,
)


def _as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_sales_header(header: Any) -> tuple[int, int] | None:
    """
    Extrage (year, month) dintr-un header "JAN 2025 - \\nJAN 2025 \\nNet Sales".
    Returnează None dacă nu e un header valid de Net Sales.
    """
    if header is None:
        return None
    text = str(header)
    if "net sales" not in text.lower():
        return None
    m = _NET_SALES_HEADER_RE.search(text)
    if m is None:
        return None
    month = _MONTH_TOKENS[m.group("mon").upper()]
    year = int(m.group("year"))
    return year, month


def _parse_sheet(ws, *, sheet_idx: int) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parsează un singur sheet. Returnează (rows, errors).
    Rândurile cu amount=0/gol sunt skipate — Sika exportă multe rânduri "ghost"
    unde produsul a existat într-o lună dar nu are vânzări în cealaltă.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    header = all_rows[0]
    if len(header) < 13:
        return [], [f"Sheet '{ws.title}': header prea scurt ({len(header)} coloane)"]

    # Detect Net Sales pairs dinamic — col 9 și 10 în layout-ul curent, dar
    # suntem defensivi: căutăm primele 2 coloane cu "Net Sales".
    ns_cols: list[tuple[int, int, int]] = []  # (col_idx, year, month)
    for col_idx, cell in enumerate(header):
        parsed = _parse_sales_header(cell)
        if parsed is not None:
            year, month = parsed
            ns_cols.append((col_idx, year, month))
        if len(ns_cols) == 2:
            break

    if len(ns_cols) < 2:
        return [], [
            f"Sheet '{ws.title}': nu am găsit 2 coloane Net Sales cu lună+an "
            f"(găsit {len(ns_cols)})"
        ]

    # Identificăm și coloanele Billed Qty asociate — pare safe să presupunem
    # aceleași offsets relative ca Net Sales (qty vine imediat după sales).
    # Layout observed: ns[0]=col9, ns[1]=col10, qty[0]=col11, qty[1]=col12.
    qty_cols: list[int] = [c + 2 for c, _, _ in ns_cols]

    # Verifică că toate cele 2 Net Sales cad în aceeași lună (diferențiate doar
    # prin an). Dacă nu, e un sheet neconvențional — avertizăm.
    if ns_cols[0][2] != ns_cols[1][2]:
        return [], [
            f"Sheet '{ws.title}': coloanele Net Sales au luni diferite "
            f"({ns_cols[0][2]} vs {ns_cols[1][2]}) — neașteptat"
        ]

    rows_out: list[dict[str, Any]] = []
    errors: list[str] = []

    for abs_idx, row in enumerate(all_rows[1:], start=2):
        if row is None or all(v is None or v == "" for v in row):
            continue
        if len(row) < 13:
            continue

        customer_name = _as_str(row[1])
        ship_to_name = _as_str(row[2])
        ship_to_code = _as_str(row[3])
        product_code = _as_str(row[5])
        product_name = _as_str(row[6])
        category_code = _as_str(row[7])

        if not customer_name or not ship_to_name:
            # Rând header duplicat sau gol la mijloc — skip în tăcere.
            continue
        if not product_code and not product_name:
            continue

        client_key = f"{customer_name} | {ship_to_name}"
        # Dacă product_code e gol, folosim product_name ca identificator
        # (aceeași politică ca la ADP).
        effective_code = product_code or product_name

        for (col_idx, year, month), qty_col in zip(ns_cols, qty_cols):
            amount_raw = row[col_idx] if col_idx < len(row) else None
            amount = _as_decimal(amount_raw)
            if amount is None or amount == 0:
                continue
            quantity = _as_decimal(row[qty_col]) if qty_col < len(row) else None

            rows_out.append({
                "year": year,
                "month": month,
                "client": client_key,
                "client_code": ship_to_code,
                "channel": "KA",
                "product_code": effective_code,
                "product_name": product_name,
                "category_code": category_code,
                "amount": amount,
                "quantity": quantity,
                "agent": None,
            })

    return rows_out, errors


def parse_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parsează toate sheet-urile cu format Sika DIY.
    Ignoră sheet-uri fără header Net Sales valid (ex. sheet-uri gol sau
    meta). Deduplică pe (year, month, client, client_code, product_code) —
    dacă același fișier listează aceleași tuple în două sheet-uri, păstrăm
    primul.
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if not wb.sheetnames:
        return [], ["Fișier Excel gol (niciun sheet)"]

    all_rows: list[dict[str, Any]] = []
    all_errors: list[str] = []
    seen: set[tuple] = set()
    sheets_parsed = 0

    for idx, ws in enumerate(wb.worksheets):
        rows, errors = _parse_sheet(ws, sheet_idx=idx)
        if errors:
            all_errors.extend(errors)
        if rows:
            sheets_parsed += 1
        for r in rows:
            key = (
                r["year"], r["month"], r["client"], r.get("client_code"),
                r.get("product_code"),
            )
            if key in seen:
                continue
            seen.add(key)
            all_rows.append(r)

    if not all_rows and not all_errors:
        return [], [
            "Niciun sheet nu conține date Sika valide. Verifică că fișierul "
            "are structura: Customer, Ship-to Party, Material, Net Sales "
            "year-1, Net Sales year."
        ]

    return all_rows, all_errors
