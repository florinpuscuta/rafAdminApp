from uuid import UUID

from fastapi import Depends, HTTPException, Request, UploadFile, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.api import APIRouter
from app.core.db import get_session
from app.modules.agents import service as agents_service
from app.modules.audit import service as audit_service
from app.modules.agents.schemas import (
    AgentAliasOut,
    AgentOut,
    AssignRequest,
    AssignmentOut,
    BulkImportResponse,
    BulkSetActiveRequest,
    BulkSetActiveResponse,
    CreateAgentAliasRequest,
    CreateAgentRequest,
    MergeAgentsRequest,
    MergeAgentsResponse,
    UnmappedAgentRow,
    UpdateAgentAliasRequest,
)
from app.modules.auth.deps import (
    get_current_admin,
    get_current_org_ids,
    get_current_tenant_id,
    get_current_user,
)
from app.modules.sales import service as sales_service
from app.modules.users.models import User

router = APIRouter(prefix="/api/agents", tags=["agents"])


@router.get("", response_model=list[AgentOut])
async def list_agents(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    agents = await agents_service.list_agents_by_tenants(session, org_ids)
    return [AgentOut.model_validate(a) for a in agents]


@router.post("", response_model=AgentOut, status_code=status.HTTP_201_CREATED)
async def create_agent(
    payload: CreateAgentRequest,
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    try:
        agent = await agents_service.create_agent(
            session,
            tenant_id=tenant_id,
            full_name=payload.full_name,
            email=payload.email,
            phone=payload.phone,
        )
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail={"code": "agent_exists", "message": "Există deja un agent cu acest nume"},
        )
    return AgentOut.model_validate(agent)


@router.get("/unmapped", response_model=list[UnmappedAgentRow])
async def list_unmapped(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    rows = await sales_service.list_agents_without_canonical_by_tenants(session, org_ids)
    return [
        UnmappedAgentRow(raw_agent=raw, row_count=count, total_amount=total)
        for raw, count, total in rows
    ]


@router.get("/aliases", response_model=list[AgentAliasOut])
async def list_aliases(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    aliases = await agents_service.list_aliases_by_tenants(session, org_ids)
    return [AgentAliasOut.model_validate(a) for a in aliases]


@router.post("/aliases", response_model=AgentAliasOut, status_code=status.HTTP_201_CREATED)
async def create_alias(
    request: Request,
    payload: CreateAgentAliasRequest,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    agent = await agents_service.get_agent(session, tenant_id, payload.agent_id)
    if agent is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "agent_not_found", "message": "Agent inexistent"},
        )

    existing = await agents_service.get_alias_by_raw(
        session, tenant_id, payload.raw_agent
    )
    if existing is not None:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail={
                "code": "alias_exists",
                "message": f"String-ul brut '{payload.raw_agent}' e deja mapat",
            },
        )

    alias = await agents_service.create_alias(
        session,
        tenant_id=tenant_id,
        raw_agent=payload.raw_agent,
        agent_id=payload.agent_id,
        resolved_by_user_id=current_user.id,
    )
    await sales_service.backfill_agent_for_raw(
        session, tenant_id, payload.raw_agent, payload.agent_id
    )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="alias.agent.created",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="agent_alias",
        target_id=alias.id,
        metadata={"raw_agent": payload.raw_agent, "agent_id": str(payload.agent_id)},
        request=request,
    )
    return AgentAliasOut.model_validate(alias)


@router.patch("/aliases/{alias_id}", response_model=AgentAliasOut)
async def update_alias(
    request: Request,
    alias_id: UUID,
    payload: UpdateAgentAliasRequest,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    alias = await agents_service.get_alias_by_id(session, tenant_id, alias_id)
    if alias is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "alias_not_found", "message": "Alias inexistent"},
        )
    new_agent = await agents_service.get_agent(session, tenant_id, payload.agent_id)
    if new_agent is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "agent_not_found", "message": "Agent nou inexistent"},
        )
    old_agent_id = alias.agent_id
    if old_agent_id == payload.agent_id:
        return AgentAliasOut.model_validate(alias)
    await sales_service.clear_agent_for_raw(session, tenant_id, alias.raw_agent)
    alias.agent_id = payload.agent_id
    alias.resolved_by_user_id = current_user.id
    from datetime import datetime, timezone
    alias.resolved_at = datetime.now(timezone.utc)
    await sales_service.backfill_agent_for_raw(
        session, tenant_id, alias.raw_agent, payload.agent_id
    )
    await session.commit()
    await session.refresh(alias)
    await audit_service.log_event(
        session,
        event_type="alias.agent.updated",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="agent_alias",
        target_id=alias.id,
        metadata={
            "raw_agent": alias.raw_agent,
            "from_agent_id": str(old_agent_id),
            "to_agent_id": str(payload.agent_id),
        },
        request=request,
    )
    return AgentAliasOut.model_validate(alias)


@router.delete("/aliases/{alias_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_alias(
    request: Request,
    alias_id: UUID,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    alias = await agents_service.get_alias_by_id(
        session, tenant_id, alias_id
    )
    if alias is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "alias_not_found", "message": "Alias inexistent"},
        )
    raw_agent = alias.raw_agent
    await sales_service.clear_agent_for_raw(session, tenant_id, raw_agent)
    await agents_service.delete_alias(session, alias)
    await audit_service.log_event(
        session,
        event_type="alias.agent.deleted",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="agent_alias",
        target_id=alias_id,
        metadata={"raw_agent": raw_agent},
        request=request,
    )
    return None


# ── Agent-Store Assignments ──────────────────────────────────────────────


@router.get("/assignments", response_model=list[AssignmentOut])
async def list_assignments(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    items = await agents_service.list_assignments_by_tenants(session, org_ids)
    return [AssignmentOut.model_validate(a) for a in items]


@router.post("/assignments", response_model=AssignmentOut, status_code=status.HTTP_201_CREATED)
async def create_assignment(
    request: Request,
    payload: AssignRequest,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    a = await agents_service.assign_store_to_agent(
        session,
        tenant_id=tenant_id,
        agent_id=payload.agent_id,
        store_id=payload.store_id,
    )
    await audit_service.log_event(
        session,
        event_type="agent_assignment.created",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="agent_assignment",
        target_id=a.id,
        metadata={"agent_id": str(payload.agent_id), "store_id": str(payload.store_id)},
        request=request,
    )
    return AssignmentOut.model_validate(a)


@router.delete("/assignments", status_code=status.HTTP_204_NO_CONTENT)
async def delete_assignment(
    request: Request,
    payload: AssignRequest,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    ok = await agents_service.unassign_store_from_agent(
        session,
        tenant_id=tenant_id,
        agent_id=payload.agent_id,
        store_id=payload.store_id,
    )
    if not ok:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "assignment_not_found", "message": "Asignare inexistentă"},
        )
    await audit_service.log_event(
        session,
        event_type="agent_assignment.deleted",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="agent_assignment",
        metadata={"agent_id": str(payload.agent_id), "store_id": str(payload.store_id)},
        request=request,
    )
    return None


@router.post("/bulk-set-active", response_model=BulkSetActiveResponse)
async def bulk_set_active(
    request: Request,
    payload: BulkSetActiveRequest,
    admin: User = Depends(get_current_admin),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Activează/dezactivează agenți multipli simultan."""
    updated = await agents_service.bulk_set_active(
        session, tenant_id=tenant_id, agent_ids=payload.ids, active=payload.active,
    )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="agent.bulk_set_active",
        tenant_id=tenant_id,
        user_id=admin.id,
        metadata={"count": updated, "active": payload.active, "ids": [str(i) for i in payload.ids]},
        request=request,
    )
    return BulkSetActiveResponse(updated=updated)


@router.post("/merge", response_model=MergeAgentsResponse)
async def merge_agents(
    request: Request,
    payload: MergeAgentsRequest,
    admin: User = Depends(get_current_admin),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    try:
        counts = await agents_service.merge_into(
            session,
            tenant_id=tenant_id,
            primary_id=payload.primary_id,
            duplicate_ids=payload.duplicate_ids,
        )
    except ValueError as exc:
        await session.rollback()
        msg = str(exc)
        raise HTTPException(
            status.HTTP_404_NOT_FOUND if msg.startswith(("primary_not_found", "duplicates_not_found")) else 400,
            detail={"code": msg.split(":")[0], "message": msg},
        )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="agent.merged",
        tenant_id=tenant_id,
        user_id=admin.id,
        target_type="agent",
        target_id=payload.primary_id,
        metadata={"duplicates": [str(d) for d in payload.duplicate_ids], **counts},
        request=request,
    )
    return MergeAgentsResponse(primary_id=payload.primary_id, **counts)


@router.post("/aliases/bulk-import", response_model=BulkImportResponse)
async def bulk_import_aliases(
    request: Request,
    file: UploadFile,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """
    Import bulk Excel cu coloane: `raw_agent`, `full_name`, `email` (opț), `phone` (opț).
    Auto-creează agenți canonici noi + alias-uri. Skip alias-urile existente.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"code": "invalid_format", "message": "Se acceptă doar .xlsx"},
        )
    content = await file.read()
    if not content:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"code": "empty_file", "message": "Fișier gol"},
        )

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return BulkImportResponse(created_agents=0, created_aliases=0, skipped=0, errors=["Excel gol"])
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip().lower().replace(" ", "_") for c in next(rows_iter)]
    except StopIteration:
        return BulkImportResponse(created_agents=0, created_aliases=0, skipped=0, errors=["Excel gol"])

    required = {"raw_agent", "full_name"}
    if not required.issubset(header):
        return BulkImportResponse(
            created_agents=0, created_aliases=0, skipped=0,
            errors=[f"Coloane obligatorii lipsă: {', '.join(required - set(header))}"],
        )
    idx = {name: i for i, name in enumerate(header)}

    existing_agents = {a.full_name: a for a in await agents_service.list_agents(session, tenant_id)}
    existing_aliases = {a.raw_agent for a in await agents_service.list_aliases(session, tenant_id)}

    created_agents = 0
    created_aliases = 0
    skipped = 0
    errors: list[str] = []

    for line_no, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None or v == "" for v in row):
            continue
        def get(col: str):
            i = idx.get(col)
            return None if i is None else row[i]

        raw_agent = str(get("raw_agent") or "").strip()
        full_name = str(get("full_name") or "").strip()
        if not raw_agent or not full_name:
            errors.append(f"Linia {line_no}: raw_agent/full_name lipsă")
            continue
        if raw_agent in existing_aliases:
            skipped += 1
            continue

        email = str(get("email") or "").strip() or None
        phone = str(get("phone") or "").strip() or None

        agent = existing_agents.get(full_name)
        if agent is None:
            try:
                agent = await agents_service.create_agent(
                    session,
                    tenant_id=tenant_id,
                    full_name=full_name,
                    email=email,
                    phone=phone,
                )
                existing_agents[full_name] = agent
                created_agents += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Linia {line_no}: {exc}")
                continue

        try:
            await agents_service.create_alias(
                session,
                tenant_id=tenant_id,
                raw_agent=raw_agent,
                agent_id=agent.id,
                resolved_by_user_id=current_user.id,
            )
            await sales_service.backfill_agent_for_raw(
                session, tenant_id, raw_agent, agent.id
            )
            existing_aliases.add(raw_agent)
            created_aliases += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Linia {line_no}: {exc}")

    await session.commit()
    await audit_service.log_event(
        session,
        event_type="alias.agent.bulk_imported",
        tenant_id=tenant_id,
        user_id=current_user.id,
        metadata={
            "filename": filename,
            "created_agents": created_agents,
            "created_aliases": created_aliases,
            "skipped": skipped,
        },
        request=request,
    )
    return BulkImportResponse(
        created_agents=created_agents,
        created_aliases=created_aliases,
        skipped=skipped,
        errors=errors[:50],
    )
