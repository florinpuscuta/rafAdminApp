"""
Ingest și rezolvare pentru StoreAgentMapping.

Fluxul:
  1. `ingest_mapping_xlsx(content, tenant_id)` — parsează fișierul de la Raf,
     upsert-ează rânduri în store_agent_mappings, creează Store și Agent
     canonic unde lipsesc, populează FK-urile.
  2. `backfill_raw_sales(session, tenant_id)` — UPDATE raw_sales SET
     store_id, agent_id FROM store_agent_mappings, join pe
     (source='ADP', client_original, ship_to_original).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.agents.models import Agent
from app.modules.mappings.models import StoreAgentMapping
from app.modules.stores.models import Store

# Headerele din fișier — exact cum le-a structurat Raf.
_EXPECTED_HEADERS = [
    "Sursa",
    "Client Original",
    "Magazin Original (ship_to)",
    "Cod Numeric Sika",
    "Agent Original",
    "Vanzari",
    "Cheie Finala (COMPLETATI)",
    "Agent Unificat (COMPLETATI)",
]


def parse_mapping_xlsx(content: bytes) -> list[dict[str, Any]]:
    """
    Parsează fișierul mapare_completa_magazine_cu_coduri_v2.xlsx.
    Ignoră rândurile fără `Cheie Finala` sau `Agent Unificat` (incomplete).
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb["Mapare Magazine"] if "Mapare Magazine" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Presupunem header pe prima linie (formatul e fix).
    header = [str(c or "").strip() for c in rows[0]]
    idx = {h: header.index(h) for h in _EXPECTED_HEADERS if h in header}

    required = {"Sursa", "Client Original", "Magazin Original (ship_to)",
                "Cheie Finala (COMPLETATI)", "Agent Unificat (COMPLETATI)"}
    if not required <= idx.keys():
        missing = required - idx.keys()
        raise ValueError(f"Headere lipsă în fișier: {missing}")

    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None:
            continue

        def get(h: str) -> Any:
            i = idx.get(h)
            return None if i is None else row[i]

        cheie = _clean(get("Cheie Finala (COMPLETATI)"))
        agent = _clean(get("Agent Unificat (COMPLETATI)"))
        source = _clean(get("Sursa"))
        client_orig = _clean(get("Client Original"))
        ship_to_orig = _clean(get("Magazin Original (ship_to)"))

        if not (cheie and agent and source and client_orig and ship_to_orig):
            continue

        out.append({
            "source": source.upper(),
            "client_original": client_orig,
            "ship_to_original": ship_to_orig,
            "agent_original": _clean(get("Agent Original")),
            "cod_numeric": _clean(get("Cod Numeric Sika")),
            "cheie_finala": cheie,
            "agent_unificat": agent,
        })
    return out


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


async def ingest_mapping_rows(
    session: AsyncSession,
    tenant_id: UUID,
    rows: list[dict[str, Any]],
) -> dict[str, int]:
    """
    Upsert în store_agent_mappings + creare canonicals (Store, Agent) pentru
    cheile noi. Idempotent — unique constraint pe (tenant, source, client,
    ship_to) previne duplicate, iar get-or-create pentru canonicals.
    """
    # Cache existing canonicals.
    existing_stores = (await session.execute(
        select(Store).where(Store.tenant_id == tenant_id)
    )).scalars().all()
    store_by_name = {s.name: s for s in existing_stores}

    existing_agents = (await session.execute(
        select(Agent).where(Agent.tenant_id == tenant_id)
    )).scalars().all()
    agent_by_name = {a.full_name: a for a in existing_agents}

    # Cache existing mappings pentru a face upsert.
    existing_mappings = (await session.execute(
        select(StoreAgentMapping).where(StoreAgentMapping.tenant_id == tenant_id)
    )).scalars().all()
    mapping_by_key = {
        (m.source, m.client_original, m.ship_to_original): m
        for m in existing_mappings
    }

    stores_created = agents_created = 0
    mappings_created = mappings_updated = 0

    for r in rows:
        cheie = r["cheie_finala"]
        agent_name = r["agent_unificat"]

        store = store_by_name.get(cheie)
        if store is None:
            chain = r["client_original"]
            city = r["ship_to_original"]
            store = Store(tenant_id=tenant_id, name=cheie, chain=chain, city=city)
            session.add(store)
            await session.flush()
            store_by_name[cheie] = store
            stores_created += 1

        agent = agent_by_name.get(agent_name)
        if agent is None:
            agent = Agent(tenant_id=tenant_id, full_name=agent_name)
            session.add(agent)
            await session.flush()
            agent_by_name[agent_name] = agent
            agents_created += 1

        key = (r["source"], r["client_original"], r["ship_to_original"])
        existing = mapping_by_key.get(key)
        if existing is None:
            session.add(StoreAgentMapping(
                tenant_id=tenant_id,
                source=r["source"],
                client_original=r["client_original"],
                ship_to_original=r["ship_to_original"],
                agent_original=r.get("agent_original"),
                cod_numeric=r.get("cod_numeric"),
                cheie_finala=cheie,
                agent_unificat=agent_name,
                store_id=store.id,
                agent_id=agent.id,
            ))
            mappings_created += 1
        else:
            if (existing.cheie_finala != cheie
                    or existing.agent_unificat != agent_name
                    or existing.store_id != store.id
                    or existing.agent_id != agent.id):
                existing.cheie_finala = cheie
                existing.agent_unificat = agent_name
                existing.store_id = store.id
                existing.agent_id = agent.id
                existing.agent_original = r.get("agent_original")
                existing.cod_numeric = r.get("cod_numeric")
                mappings_updated += 1

    await session.flush()
    return {
        "rows_processed": len(rows),
        "stores_created": stores_created,
        "agents_created": agents_created,
        "mappings_created": mappings_created,
        "mappings_updated": mappings_updated,
    }


async def backfill_raw_sales(
    session: AsyncSession,
    tenant_id: UUID,
    *,
    source: str = "ADP",
) -> dict[str, int]:
    """
    UPDATE raw_sales SET store_id, agent_id folosind store_agent_mappings.

    ADP: match pe `client = "{client_original} | {ship_to_original}"`
      (formatul combined_key din importer).

    SIKA: match dual — PRIMAR pe `client_code = m.cod_numeric` (ship-to numeric,
      stabil între exporturi), apoi FALLBACK pe nume (același format combined)
      pentru rândurile rămase nemapate. Codul e populat pe ~203/209 mapări SIKA,
      fallback-ul prinde restul.

    Scope de ștergere înainte: DOAR rândurile sursei respective — identificăm
    prin import_batches.source ('sales_xlsx' = ADP, 'sika_xlsx' = SIKA). Astfel
    un reload SIKA nu șterge mapările ADP și vice-versa.
    """
    src_upper = source.upper()
    batch_source = "sika_xlsx" if src_upper == "SIKA" else "sales_xlsx"

    # Golim întâi ce am setat anterior pe rândurile DIN aceeași sursă — altfel
    # dacă un store/agent a fost redenumit în fișierul Raf, vechea mapare rămâne.
    await session.execute(text("""
        UPDATE raw_sales rs
        SET store_id = NULL, agent_id = NULL
        FROM import_batches b
        WHERE b.id = rs.batch_id
          AND b.source = :bsrc
          AND rs.tenant_id = :tid
          AND UPPER(rs.channel) = 'KA'
    """), {"tid": str(tenant_id), "bsrc": batch_source})

    if src_upper == "SIKA":
        # Primar: cod ship-to.
        primary = text("""
            UPDATE raw_sales rs
            SET store_id = m.store_id,
                agent_id = m.agent_id
            FROM store_agent_mappings m, import_batches b
            WHERE b.id = rs.batch_id
              AND b.source = :bsrc
              AND m.tenant_id = rs.tenant_id
              AND m.source = 'SIKA'
              AND m.cod_numeric IS NOT NULL
              AND rs.client_code = m.cod_numeric
              AND rs.tenant_id = :tid
              AND UPPER(rs.channel) = 'KA'
              AND rs.store_id IS NULL
        """)
        rc_primary = (await session.execute(
            primary, {"tid": str(tenant_id), "bsrc": batch_source}
        )).rowcount or 0

        # Fallback: nume combined.
        fallback = text("""
            UPDATE raw_sales rs
            SET store_id = m.store_id,
                agent_id = m.agent_id
            FROM store_agent_mappings m, import_batches b
            WHERE b.id = rs.batch_id
              AND b.source = :bsrc
              AND m.tenant_id = rs.tenant_id
              AND m.source = 'SIKA'
              AND rs.client = m.client_original || ' | ' || m.ship_to_original
              AND rs.tenant_id = :tid
              AND UPPER(rs.channel) = 'KA'
              AND rs.store_id IS NULL
        """)
        rc_fallback = (await session.execute(
            fallback, {"tid": str(tenant_id), "bsrc": batch_source}
        )).rowcount or 0
        return {
            "rows_updated": rc_primary + rc_fallback,
            "rows_by_code": rc_primary,
            "rows_by_name": rc_fallback,
        }

    # ADP: match pe nume combined.
    stmt = text("""
        UPDATE raw_sales rs
        SET store_id = m.store_id,
            agent_id = m.agent_id
        FROM store_agent_mappings m, import_batches b
        WHERE b.id = rs.batch_id
          AND b.source = :bsrc
          AND m.tenant_id = rs.tenant_id
          AND m.source = :src
          AND rs.client = m.client_original || ' | ' || m.ship_to_original
          AND rs.tenant_id = :tid
          AND UPPER(rs.channel) = 'KA'
    """)
    result = await session.execute(
        stmt, {"tid": str(tenant_id), "src": src_upper, "bsrc": batch_source}
    )
    return {"rows_updated": result.rowcount or 0}


async def list_mappings(
    session: AsyncSession, tenant_id: UUID, *, source: str | None = None,
) -> list[StoreAgentMapping]:
    q = select(StoreAgentMapping).where(StoreAgentMapping.tenant_id == tenant_id)
    if source:
        q = q.where(StoreAgentMapping.source == source.upper())
    q = q.order_by(StoreAgentMapping.cheie_finala, StoreAgentMapping.source)
    return list((await session.execute(q)).scalars().all())


async def _get_or_create_store(
    session: AsyncSession, tenant_id: UUID, cheie: str,
    chain: str, city: str,
) -> Store:
    existing = (await session.execute(
        select(Store).where(Store.tenant_id == tenant_id, Store.name == cheie)
    )).scalar_one_or_none()
    if existing is not None:
        return existing
    s = Store(tenant_id=tenant_id, name=cheie, chain=chain, city=city)
    session.add(s)
    await session.flush()
    return s


async def _get_or_create_agent(
    session: AsyncSession, tenant_id: UUID, full_name: str,
) -> Agent:
    existing = (await session.execute(
        select(Agent).where(
            Agent.tenant_id == tenant_id, Agent.full_name == full_name
        )
    )).scalar_one_or_none()
    if existing is not None:
        return existing
    a = Agent(tenant_id=tenant_id, full_name=full_name)
    session.add(a)
    await session.flush()
    return a


async def get_mapping(
    session: AsyncSession, tenant_id: UUID, mapping_id: UUID,
) -> StoreAgentMapping | None:
    return (await session.execute(
        select(StoreAgentMapping).where(
            StoreAgentMapping.tenant_id == tenant_id,
            StoreAgentMapping.id == mapping_id,
        )
    )).scalar_one_or_none()


async def create_mapping(
    session: AsyncSession, tenant_id: UUID, data: dict[str, Any],
) -> StoreAgentMapping:
    source = data["source"].upper().strip()
    client_original = data["client_original"].strip()
    ship_to_original = data["ship_to_original"].strip()
    cheie = data["cheie_finala"].strip()
    agent_name = data["agent_unificat"].strip()

    store = await _get_or_create_store(
        session, tenant_id, cheie, client_original, ship_to_original,
    )
    agent = await _get_or_create_agent(session, tenant_id, agent_name)

    m = StoreAgentMapping(
        tenant_id=tenant_id,
        source=source,
        client_original=client_original,
        ship_to_original=ship_to_original,
        agent_original=(data.get("agent_original") or None),
        cod_numeric=(data.get("cod_numeric") or None),
        cheie_finala=cheie,
        agent_unificat=agent_name,
        store_id=store.id,
        agent_id=agent.id,
    )
    session.add(m)
    await session.flush()
    return m


async def update_mapping(
    session: AsyncSession, tenant_id: UUID, mapping_id: UUID,
    data: dict[str, Any],
) -> StoreAgentMapping | None:
    m = await get_mapping(session, tenant_id, mapping_id)
    if m is None:
        return None

    if "source" in data and data["source"] is not None:
        m.source = data["source"].upper().strip()
    if "client_original" in data and data["client_original"] is not None:
        m.client_original = data["client_original"].strip()
    if "ship_to_original" in data and data["ship_to_original"] is not None:
        m.ship_to_original = data["ship_to_original"].strip()
    if "agent_original" in data:
        m.agent_original = (data["agent_original"] or None)
    if "cod_numeric" in data:
        m.cod_numeric = (data["cod_numeric"] or None)

    if "cheie_finala" in data and data["cheie_finala"] is not None:
        cheie = data["cheie_finala"].strip()
        m.cheie_finala = cheie
        store = await _get_or_create_store(
            session, tenant_id, cheie, m.client_original, m.ship_to_original,
        )
        m.store_id = store.id
    if "agent_unificat" in data and data["agent_unificat"] is not None:
        agent_name = data["agent_unificat"].strip()
        m.agent_unificat = agent_name
        agent = await _get_or_create_agent(session, tenant_id, agent_name)
        m.agent_id = agent.id

    await session.flush()
    return m


async def delete_mapping(
    session: AsyncSession, tenant_id: UUID, mapping_id: UUID,
) -> bool:
    m = await get_mapping(session, tenant_id, mapping_id)
    if m is None:
        return False
    await session.delete(m)
    await session.flush()
    return True


# ── Magazine nealocate ─────────────────────────────────────────────────────

_UNMAPPED_SCOPE_SOURCES: dict[str, tuple[list[str], str]] = {
    # scope → (batch_sources, SAM.source corespunzător)
    "adp": (["sales_xlsx"], "ADP"),
    "sika": (["sika_mtd_xlsx", "sika_xlsx"], "SIKA"),
}


async def list_unmapped_clients(
    session: AsyncSession, tenant_id: UUID, *, scope: str,
) -> list[dict[str, Any]]:
    """Distinct (client, ship_to) din raw_sales KA care NU sunt acoperite de
    o intrare SAM cu agent_id non-NULL. Returnează volumele agregate ca
    reper vizual pentru alocare.
    """
    cfg = _UNMAPPED_SCOPE_SOURCES.get(scope.lower())
    if cfg is None:
        return []
    batch_sources, sam_source = cfg

    stmt = text("""
        SELECT
          SPLIT_PART(rs.client, ' | ', 1) AS client_original,
          SUBSTRING(rs.client FROM POSITION(' | ' IN rs.client) + 3) AS ship_to_original,
          rs.client AS raw_client,
          COUNT(*) AS row_count,
          COALESCE(SUM(rs.amount), 0) AS total_sales
        FROM raw_sales rs
        JOIN import_batches b ON b.id = rs.batch_id
        WHERE rs.tenant_id = :tid
          AND UPPER(rs.channel) = 'KA'
          AND b.source = ANY(:sources)
          AND POSITION(' | ' IN rs.client) > 0
          AND NOT EXISTS (
            SELECT 1 FROM store_agent_mappings sam
            WHERE sam.tenant_id = rs.tenant_id
              AND sam.agent_id IS NOT NULL
              AND UPPER(TRIM(sam.client_original || ' | ' || sam.ship_to_original))
                  = UPPER(TRIM(rs.client))
          )
        GROUP BY rs.client
        ORDER BY total_sales DESC
    """)
    result = await session.execute(
        stmt, {"tid": str(tenant_id), "sources": batch_sources},
    )
    return [
        {
            "client_original": r.client_original.strip(),
            "ship_to_original": r.ship_to_original.strip(),
            "raw_client": r.raw_client,
            "row_count": int(r.row_count),
            "total_sales": r.total_sales,
            "source": sam_source,
        }
        for r in result.all()
    ]
