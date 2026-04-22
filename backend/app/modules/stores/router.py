from uuid import UUID

from fastapi import Depends, HTTPException, Request, UploadFile, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.api import APIRouter
from app.core.db import get_session
from app.modules.audit import service as audit_service
from app.modules.auth.deps import get_current_admin, get_current_tenant_id, get_current_user
from app.modules.sales import service as sales_service
from app.modules.stores import service as stores_service
from app.modules.stores.schemas import (
    BulkImportResponse,
    BulkSetActiveRequest,
    BulkSetActiveResponse,
    CreateAliasRequest,
    CreateStoreRequest,
    MergeStoresRequest,
    MergeStoresResponse,
    StoreAliasOut,
    StoreOut,
    SuggestedMatch,
    SuggestionRow,
    UnmappedClientRow,
    UpdateAliasRequest,
)
from app.modules.users.models import User

router = APIRouter(prefix="/api/stores", tags=["stores"])


@router.get("", response_model=list[StoreOut])
async def list_stores(
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    stores = await stores_service.list_stores(session, tenant_id)
    return [StoreOut.model_validate(s) for s in stores]


@router.get("/chains", response_model=list[str])
async def list_chains(
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    return await stores_service.list_chains(session, tenant_id)


@router.post("", response_model=StoreOut, status_code=status.HTTP_201_CREATED)
async def create_store(
    payload: CreateStoreRequest,
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    try:
        store = await stores_service.create_store(
            session,
            tenant_id=tenant_id,
            name=payload.name,
            chain=payload.chain,
            city=payload.city,
        )
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail={"code": "store_exists", "message": "Există deja un magazin cu acest nume"},
        )
    return StoreOut.model_validate(store)


@router.get("/unmapped", response_model=list[UnmappedClientRow])
async def list_unmapped(
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    rows = await sales_service.list_clients_without_store(session, tenant_id)
    return [
        UnmappedClientRow(raw_client=client, row_count=count, total_amount=total)
        for client, count, total in rows
    ]


@router.get("/unmapped/suggestions", response_model=list[SuggestionRow])
async def unmapped_suggestions(
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Sugestii fuzzy-match pentru fiecare raw_client nemapat."""
    rows = await sales_service.list_clients_without_store(session, tenant_id)
    raw_clients = [r[0] for r in rows]
    stores = await stores_service.list_stores(session, tenant_id)
    if not stores:
        return [SuggestionRow(raw_client=r, suggestions=[]) for r in raw_clients]
    matches = stores_service.suggest_matches(raw_clients, stores)
    return [
        SuggestionRow(
            raw_client=r,
            suggestions=[
                SuggestedMatch(store_id=sid, store_name=name, score=score)
                for sid, name, score in matches.get(r, [])
            ],
        )
        for r in raw_clients
    ]


@router.get("/aliases", response_model=list[StoreAliasOut])
async def list_aliases(
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    aliases = await stores_service.list_aliases(session, tenant_id)
    return [StoreAliasOut.model_validate(a) for a in aliases]


@router.post("/aliases", response_model=StoreAliasOut, status_code=status.HTTP_201_CREATED)
async def create_alias(
    request: Request,
    payload: CreateAliasRequest,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    store = await stores_service.get_store(session, current_user.tenant_id, payload.store_id)
    if store is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "store_not_found", "message": "Magazin inexistent"},
        )

    existing = await stores_service.get_alias_by_raw(
        session, current_user.tenant_id, payload.raw_client
    )
    if existing is not None:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail={
                "code": "alias_exists",
                "message": f"String-ul brut '{payload.raw_client}' e deja mapat",
            },
        )

    alias = await stores_service.create_alias(
        session,
        tenant_id=current_user.tenant_id,
        raw_client=payload.raw_client,
        store_id=payload.store_id,
        resolved_by_user_id=current_user.id,
    )
    await sales_service.backfill_store_for_client(
        session, current_user.tenant_id, payload.raw_client, payload.store_id
    )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="alias.store.created",
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        target_type="store_alias",
        target_id=alias.id,
        metadata={"raw_client": payload.raw_client, "store_id": str(payload.store_id)},
        request=request,
    )
    return StoreAliasOut.model_validate(alias)


@router.patch("/aliases/{alias_id}", response_model=StoreAliasOut)
async def update_alias(
    request: Request,
    alias_id: UUID,
    payload: UpdateAliasRequest,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    alias = await stores_service.get_alias_by_id(session, current_user.tenant_id, alias_id)
    if alias is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "alias_not_found", "message": "Alias inexistent"},
        )
    new_store = await stores_service.get_store(session, current_user.tenant_id, payload.store_id)
    if new_store is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "store_not_found", "message": "Magazin nou inexistent"},
        )
    old_store_id = alias.store_id
    if old_store_id == payload.store_id:
        return StoreAliasOut.model_validate(alias)

    # 1) clear pe raw_sales vechi (pentru acest raw_client)
    await sales_service.clear_store_for_client(
        session, current_user.tenant_id, alias.raw_client
    )
    # 2) modifică alias
    alias.store_id = payload.store_id
    alias.resolved_by_user_id = current_user.id
    from datetime import datetime, timezone
    alias.resolved_at = datetime.now(timezone.utc)
    # 3) backfill cu noul store
    await sales_service.backfill_store_for_client(
        session, current_user.tenant_id, alias.raw_client, payload.store_id
    )
    await session.commit()
    await session.refresh(alias)
    await audit_service.log_event(
        session,
        event_type="alias.store.updated",
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        target_type="store_alias",
        target_id=alias.id,
        metadata={
            "raw_client": alias.raw_client,
            "from_store_id": str(old_store_id),
            "to_store_id": str(payload.store_id),
        },
        request=request,
    )
    return StoreAliasOut.model_validate(alias)


@router.delete("/aliases/{alias_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_alias(
    request: Request,
    alias_id: UUID,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    alias = await stores_service.get_alias_by_id(
        session, current_user.tenant_id, alias_id
    )
    if alias is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "alias_not_found", "message": "Alias inexistent"},
        )
    raw_client = alias.raw_client
    await sales_service.clear_store_for_client(
        session, current_user.tenant_id, raw_client
    )
    await stores_service.delete_alias(session, alias)
    await audit_service.log_event(
        session,
        event_type="alias.store.deleted",
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        target_type="store_alias",
        target_id=alias_id,
        metadata={"raw_client": raw_client},
        request=request,
    )
    return None


@router.post("/bulk-set-active", response_model=BulkSetActiveResponse)
async def bulk_set_active(
    request: Request,
    payload: BulkSetActiveRequest,
    admin: User = Depends(get_current_admin),
    session: AsyncSession = Depends(get_session),
):
    """Activează/dezactivează magazine multiple simultan."""
    updated = await stores_service.bulk_set_active(
        session, tenant_id=admin.tenant_id, store_ids=payload.ids, active=payload.active,
    )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="store.bulk_set_active",
        tenant_id=admin.tenant_id,
        user_id=admin.id,
        metadata={"count": updated, "active": payload.active, "ids": [str(i) for i in payload.ids]},
        request=request,
    )
    return BulkSetActiveResponse(updated=updated)


@router.post("/merge", response_model=MergeStoresResponse)
async def merge_stores(
    request: Request,
    payload: MergeStoresRequest,
    admin: User = Depends(get_current_admin),
    session: AsyncSession = Depends(get_session),
):
    """Consolidează duplicate-ids în primary-id (transfer alias + raw_sales + assignments)."""
    try:
        counts = await stores_service.merge_into(
            session,
            tenant_id=admin.tenant_id,
            primary_id=payload.primary_id,
            duplicate_ids=payload.duplicate_ids,
        )
    except ValueError as exc:
        await session.rollback()
        msg = str(exc)
        raise HTTPException(
            status.HTTP_404_NOT_FOUND if msg.startswith(("primary_not_found", "duplicates_not_found")) else 400,
            detail={"code": msg.split(":")[0], "message": msg},
        )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="store.merged",
        tenant_id=admin.tenant_id,
        user_id=admin.id,
        target_type="store",
        target_id=payload.primary_id,
        metadata={"duplicates": [str(d) for d in payload.duplicate_ids], **counts},
        request=request,
    )
    return MergeStoresResponse(primary_id=payload.primary_id, **counts)


@router.post("/aliases/bulk-import", response_model=BulkImportResponse)
async def bulk_import_aliases(
    request: Request,
    file: UploadFile,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """
    Import bulk din Excel cu coloane: `raw_client`, `store_name`, `chain` (opț), `city` (opț).
    Auto-creează magazine canonice care nu există + creează alias-urile.
    Skip linii cu alias deja mapat.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"code": "invalid_format", "message": "Se acceptă doar .xlsx"},
        )
    content = await file.read()
    if not content:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"code": "empty_file", "message": "Fișier gol"},
        )

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return BulkImportResponse(created_stores=0, created_aliases=0, skipped=0, errors=["Excel gol"])
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip().lower().replace(" ", "_") for c in next(rows_iter)]
    except StopIteration:
        return BulkImportResponse(created_stores=0, created_aliases=0, skipped=0, errors=["Excel gol"])

    required = {"raw_client", "store_name"}
    if not required.issubset(header):
        return BulkImportResponse(
            created_stores=0, created_aliases=0, skipped=0,
            errors=[f"Coloane obligatorii lipsă: {', '.join(required - set(header))}"],
        )
    idx = {name: i for i, name in enumerate(header)}

    # Preîncărcăm magazinele existente + alias-urile existente
    existing_stores = {s.name: s for s in await stores_service.list_stores(session, current_user.tenant_id)}
    existing_aliases = {a.raw_client for a in await stores_service.list_aliases(session, current_user.tenant_id)}

    created_stores = 0
    created_aliases = 0
    skipped = 0
    errors: list[str] = []

    for line_no, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None or v == "" for v in row):
            continue
        def get(col: str):
            i = idx.get(col)
            return None if i is None else row[i]

        raw_client = str(get("raw_client") or "").strip()
        store_name = str(get("store_name") or "").strip()
        if not raw_client or not store_name:
            errors.append(f"Linia {line_no}: raw_client/store_name lipsă")
            continue
        if raw_client in existing_aliases:
            skipped += 1
            continue

        chain = str(get("chain") or "").strip() or None
        city = str(get("city") or "").strip() or None

        store = existing_stores.get(store_name)
        if store is None:
            try:
                store = await stores_service.create_store(
                    session,
                    tenant_id=current_user.tenant_id,
                    name=store_name,
                    chain=chain,
                    city=city,
                )
                existing_stores[store_name] = store
                created_stores += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Linia {line_no}: {exc}")
                continue

        try:
            await stores_service.create_alias(
                session,
                tenant_id=current_user.tenant_id,
                raw_client=raw_client,
                store_id=store.id,
                resolved_by_user_id=current_user.id,
            )
            await sales_service.backfill_store_for_client(
                session, current_user.tenant_id, raw_client, store.id
            )
            existing_aliases.add(raw_client)
            created_aliases += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Linia {line_no}: {exc}")

    await session.commit()
    await audit_service.log_event(
        session,
        event_type="alias.store.bulk_imported",
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        metadata={
            "filename": filename,
            "created_stores": created_stores,
            "created_aliases": created_aliases,
            "skipped": skipped,
        },
        request=request,
    )
    return BulkImportResponse(
        created_stores=created_stores,
        created_aliases=created_aliases,
        skipped=skipped,
        errors=errors[:50],
    )
