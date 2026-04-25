from uuid import UUID

from fastapi import Depends, HTTPException, Request, UploadFile, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.api import APIRouter
from app.core.db import get_session
from app.modules.audit import service as audit_service
from app.modules.auth.deps import (
    get_current_admin,
    get_current_org_ids,
    get_current_tenant_id,
    get_current_user,
)
from app.modules.products import service as products_service
from app.modules.products.schemas import (
    BulkImportResponse,
    BulkSetActiveRequest,
    BulkSetActiveResponse,
    CreateProductAliasRequest,
    CreateProductRequest,
    MergeProductsRequest,
    MergeProductsResponse,
    ProductAliasOut,
    ProductOut,
    UnmappedProductRow,
    UpdateProductAliasRequest,
)
from app.modules.sales import service as sales_service
from app.modules.users.models import User

router = APIRouter(prefix="/api/products", tags=["products"])


@router.get("", response_model=list[ProductOut])
async def list_products(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    products = await products_service.list_products_by_tenants(session, org_ids)
    return [ProductOut.model_validate(p) for p in products]


@router.get("/categories", response_model=list[str])
async def list_categories(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    return await products_service.list_categories_by_tenants(session, org_ids)


@router.post("", response_model=ProductOut, status_code=status.HTTP_201_CREATED)
async def create_product(
    payload: CreateProductRequest,
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    try:
        product = await products_service.create_product(
            session,
            tenant_id=tenant_id,
            code=payload.code,
            name=payload.name,
            category=payload.category,
            brand=payload.brand,
        )
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail={"code": "product_exists", "message": "Există deja un produs cu acest cod"},
        )
    return ProductOut.model_validate(product)


@router.get("/unmapped", response_model=list[UnmappedProductRow])
async def list_unmapped(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    rows = await sales_service.list_products_without_canonical_by_tenants(session, org_ids)
    return [
        UnmappedProductRow(
            raw_code=code, sample_name=name, row_count=count, total_amount=total
        )
        for code, name, count, total in rows
    ]


@router.get("/aliases", response_model=list[ProductAliasOut])
async def list_aliases(
    org_ids: list[UUID] = Depends(get_current_org_ids),
    session: AsyncSession = Depends(get_session),
):
    aliases = await products_service.list_aliases_by_tenants(session, org_ids)
    return [ProductAliasOut.model_validate(a) for a in aliases]


@router.post("/aliases", response_model=ProductAliasOut, status_code=status.HTTP_201_CREATED)
async def create_alias(
    request: Request,
    payload: CreateProductAliasRequest,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    product = await products_service.get_product(
        session, tenant_id, payload.product_id
    )
    if product is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "product_not_found", "message": "Produs inexistent"},
        )

    existing = await products_service.get_alias_by_raw(
        session, tenant_id, payload.raw_code
    )
    if existing is not None:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail={
                "code": "alias_exists",
                "message": f"Codul '{payload.raw_code}' e deja mapat",
            },
        )

    alias = await products_service.create_alias(
        session,
        tenant_id=tenant_id,
        raw_code=payload.raw_code,
        product_id=payload.product_id,
        resolved_by_user_id=current_user.id,
    )
    await sales_service.backfill_product_for_raw(
        session, tenant_id, payload.raw_code, payload.product_id
    )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="alias.product.created",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="product_alias",
        target_id=alias.id,
        metadata={"raw_code": payload.raw_code, "product_id": str(payload.product_id)},
        request=request,
    )
    return ProductAliasOut.model_validate(alias)


@router.patch("/aliases/{alias_id}", response_model=ProductAliasOut)
async def update_alias(
    request: Request,
    alias_id: UUID,
    payload: UpdateProductAliasRequest,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    alias = await products_service.get_alias_by_id(session, tenant_id, alias_id)
    if alias is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "alias_not_found", "message": "Alias inexistent"},
        )
    new_product = await products_service.get_product(
        session, tenant_id, payload.product_id
    )
    if new_product is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "product_not_found", "message": "Produs nou inexistent"},
        )
    old_product_id = alias.product_id
    if old_product_id == payload.product_id:
        return ProductAliasOut.model_validate(alias)
    await sales_service.clear_product_for_raw(
        session, tenant_id, alias.raw_code
    )
    alias.product_id = payload.product_id
    alias.resolved_by_user_id = current_user.id
    from datetime import datetime, timezone
    alias.resolved_at = datetime.now(timezone.utc)
    await sales_service.backfill_product_for_raw(
        session, tenant_id, alias.raw_code, payload.product_id
    )
    await session.commit()
    await session.refresh(alias)
    await audit_service.log_event(
        session,
        event_type="alias.product.updated",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="product_alias",
        target_id=alias.id,
        metadata={
            "raw_code": alias.raw_code,
            "from_product_id": str(old_product_id),
            "to_product_id": str(payload.product_id),
        },
        request=request,
    )
    return ProductAliasOut.model_validate(alias)


@router.delete("/aliases/{alias_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_alias(
    request: Request,
    alias_id: UUID,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    alias = await products_service.get_alias_by_id(
        session, tenant_id, alias_id
    )
    if alias is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"code": "alias_not_found", "message": "Alias inexistent"},
        )
    raw_code = alias.raw_code
    await sales_service.clear_product_for_raw(session, tenant_id, raw_code)
    await products_service.delete_alias(session, alias)
    await audit_service.log_event(
        session,
        event_type="alias.product.deleted",
        tenant_id=tenant_id,
        user_id=current_user.id,
        target_type="product_alias",
        target_id=alias_id,
        metadata={"raw_code": raw_code},
        request=request,
    )
    return None


@router.post("/bulk-set-active", response_model=BulkSetActiveResponse)
async def bulk_set_active(
    request: Request,
    payload: BulkSetActiveRequest,
    admin: User = Depends(get_current_admin),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """Activează/dezactivează produse multiple simultan."""
    updated = await products_service.bulk_set_active(
        session, tenant_id=tenant_id, product_ids=payload.ids, active=payload.active,
    )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="product.bulk_set_active",
        tenant_id=tenant_id,
        user_id=admin.id,
        metadata={"count": updated, "active": payload.active, "ids": [str(i) for i in payload.ids]},
        request=request,
    )
    return BulkSetActiveResponse(updated=updated)


@router.post("/merge", response_model=MergeProductsResponse)
async def merge_products(
    request: Request,
    payload: MergeProductsRequest,
    admin: User = Depends(get_current_admin),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    try:
        counts = await products_service.merge_into(
            session,
            tenant_id=tenant_id,
            primary_id=payload.primary_id,
            duplicate_ids=payload.duplicate_ids,
        )
    except ValueError as exc:
        await session.rollback()
        msg = str(exc)
        raise HTTPException(
            status.HTTP_404_NOT_FOUND if msg.startswith(("primary_not_found", "duplicates_not_found")) else 400,
            detail={"code": msg.split(":")[0], "message": msg},
        )
    await session.commit()
    await audit_service.log_event(
        session,
        event_type="product.merged",
        tenant_id=tenant_id,
        user_id=admin.id,
        target_type="product",
        target_id=payload.primary_id,
        metadata={"duplicates": [str(d) for d in payload.duplicate_ids], **counts},
        request=request,
    )
    return MergeProductsResponse(primary_id=payload.primary_id, **counts)


@router.post("/aliases/bulk-import", response_model=BulkImportResponse)
async def bulk_import_aliases(
    request: Request,
    file: UploadFile,
    current_user: User = Depends(get_current_user),
    tenant_id: UUID = Depends(get_current_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    """
    Excel cu coloane: `raw_code`, `code`, `name`, `category` (opț), `brand` (opț).
    Auto-creează produse canonice noi + alias-urile.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"code": "invalid_format", "message": "Se acceptă doar .xlsx"},
        )
    content = await file.read()
    if not content:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"code": "empty_file", "message": "Fișier gol"},
        )

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return BulkImportResponse(created_products=0, created_aliases=0, skipped=0, errors=["Excel gol"])
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip().lower().replace(" ", "_") for c in next(rows_iter)]
    except StopIteration:
        return BulkImportResponse(created_products=0, created_aliases=0, skipped=0, errors=["Excel gol"])

    required = {"raw_code", "code", "name"}
    if not required.issubset(header):
        return BulkImportResponse(
            created_products=0, created_aliases=0, skipped=0,
            errors=[f"Coloane obligatorii lipsă: {', '.join(required - set(header))}"],
        )
    idx = {name: i for i, name in enumerate(header)}

    existing_products = {p.code: p for p in await products_service.list_products(session, tenant_id)}
    existing_aliases = {a.raw_code for a in await products_service.list_aliases(session, tenant_id)}

    created_products = 0
    created_aliases = 0
    skipped = 0
    errors: list[str] = []

    for line_no, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None or v == "" for v in row):
            continue
        def get(col: str):
            i = idx.get(col)
            return None if i is None else row[i]

        raw_code = str(get("raw_code") or "").strip()
        code = str(get("code") or "").strip()
        name = str(get("name") or "").strip()
        if not raw_code or not code or not name:
            errors.append(f"Linia {line_no}: raw_code/code/name lipsă")
            continue
        if raw_code in existing_aliases:
            skipped += 1
            continue

        category = str(get("category") or "").strip() or None
        brand = str(get("brand") or "").strip() or None

        product = existing_products.get(code)
        if product is None:
            try:
                product = await products_service.create_product(
                    session,
                    tenant_id=tenant_id,
                    code=code,
                    name=name,
                    category=category,
                    brand=brand,
                )
                existing_products[code] = product
                created_products += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Linia {line_no}: {exc}")
                continue

        try:
            await products_service.create_alias(
                session,
                tenant_id=tenant_id,
                raw_code=raw_code,
                product_id=product.id,
                resolved_by_user_id=current_user.id,
            )
            await sales_service.backfill_product_for_raw(
                session, tenant_id, raw_code, product.id
            )
            existing_aliases.add(raw_code)
            created_aliases += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Linia {line_no}: {exc}")

    await session.commit()
    await audit_service.log_event(
        session,
        event_type="alias.product.bulk_imported",
        tenant_id=tenant_id,
        user_id=current_user.id,
        metadata={
            "filename": filename,
            "created_products": created_products,
            "created_aliases": created_aliases,
            "skipped": skipped,
        },
        request=request,
    )
    return BulkImportResponse(
        created_products=created_products,
        created_aliases=created_aliases,
        skipped=skipped,
        errors=errors[:50],
    )
