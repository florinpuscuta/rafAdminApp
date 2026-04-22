"""
Parser Excel → RawOrder pentru upload-ul Sika (comenzi open).

Format: un singur sheet cu comenzi open, fără split pe status. Toate
rândurile primesc status='OPEN'.

Layout coloane (0-indexat, preluat din `process_sika_exercitiu`):
   1  client (nume)
   2  ship_to_code (cod numeric)
   3  ship_to_name
   4  target_market            → category_code
   5  product_code
   6  product_name
   8  open_qty                 → quantity
   9  open_amount              → amount
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def _as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def parse_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parsează primul sheet. Întoarce (rows, errors).
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if not wb.sheetnames:
        return [], ["Fișier Excel gol (niciun sheet)"]

    ws = wb[wb.sheetnames[0]]
    rows_out: list[dict[str, Any]] = []
    errors: list[str] = []

    for abs_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or v == "" for v in row):
            continue
        if len(row) < 10:
            continue

        client_name = _as_str(row[1])
        ship_to_code = _as_str(row[2])
        ship_to_name = _as_str(row[3])
        category_code = _as_str(row[4])
        product_code = _as_str(row[5])
        product_name = _as_str(row[6])
        quantity = _as_decimal(row[8])
        amount = _as_decimal(row[9])

        if not client_name:
            continue
        if amount is None or amount == 0:
            continue
        if not product_code and not product_name:
            continue

        ship_to = ship_to_name or ship_to_code
        client_key = f"{client_name} | {ship_to}" if ship_to else client_name
        effective_code = product_code or product_name

        rows_out.append({
            "source": "sika",
            "chain": None,
            "client": client_key,
            "client_code": ship_to_code,
            "ship_to": ship_to,
            "nr_comanda": None,
            "product_code": effective_code,
            "product_name": product_name,
            "category_code": category_code,
            "status": "OPEN",
            "amount": amount,
            "quantity": quantity,
            "remaining_amount": amount,
            "remaining_quantity": quantity,
            "data_livrare": None,
            "ind": None,
            "has_ind": False,
        })

    if not rows_out:
        errors.append(
            "Niciun rând valid. Verifică format: col 2=Client, col 3=Ship-to cod, "
            "col 4=Ship-to nume, col 5=Target Market, col 6=Material cod, "
            "col 7=Material denumire, col 9=Open Qty, col 10=Open Amount."
        )

    return rows_out, errors
