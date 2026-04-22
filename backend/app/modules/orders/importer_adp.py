"""
Parser Excel → RawOrder pentru upload-ul ADP (radComenzi).

Format: un sheet per lanț KA (Dedeman/Altex/Leroy Merlin/Hornbach). Per sheet
se iterează liniile; fiecare rând produce un RawOrder cu status NELIVRAT sau
NEFACTURAT (NEFACTURARE se normalizează la NEFACTURAT).

Layout coloane (0-indexat, identic cu legacy `radiography_service.py`):
   1  nr_comanda
   4  cod_articol            → product_code
   5  descriere              → product_name
   8  cantitate              → quantity
  11  valoare                → amount (suma completă, nu proporțional)
  12  data_livrare (string)
  13  cantitate_rest         → remaining_quantity
  14  status (NELIVRAT / NEFACTURAT / NEFACTURARE)
  15  tip2                   → category_code
  16  ind
  17  ship_to (nume)

NELIVRAT skipat dacă cant_rest ≤ 0. NEFACTURAT include mereu (poate avea 0).
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


# ── Map nume sheet → chain canonic + client full name ─────────────────────
# Preluat din utils/constants.py din legacy. Sheet-uri care NU sunt în această
# structură sunt ignorate (de obicei rapoarte agregate / TOTAL).
_CHAIN_MAP: dict[str, tuple[str, str]] = {
    "Dedeman":      ("Dedeman", "DEDEMAN SRL"),
    "Altex":        ("Altex", "ALTEX ROMANIA SRL"),
    "Leroy Merlin": ("Leroy Merlin", "LEROY MERLIN ROMANIA SRL"),
    "Hornbach":     ("Hornbach", "HORNBACH CENTRALA SRL"),
    "DEDEMAN":      ("Dedeman", "DEDEMAN SRL"),
    "ALTEX":        ("Altex", "ALTEX ROMANIA SRL"),
    "LEROY":        ("Leroy Merlin", "LEROY MERLIN ROMANIA SRL"),
    "HORNBACH":     ("Hornbach", "HORNBACH CENTRALA SRL"),
}


def _as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def parse_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parsează toate sheet-urile KA și întoarce (rows, errors).

    Fiecare rând e un dict cu cheile necesare pentru RawOrder — fără
    tenant_id/batch_id/report_date (se adaugă la insert).
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if not wb.sheetnames:
        return [], ["Fișier Excel gol (niciun sheet)"]

    rows_out: list[dict[str, Any]] = []
    errors: list[str] = []
    sheets_parsed = 0

    for ws in wb.worksheets:
        mapping = _CHAIN_MAP.get(ws.title.strip())
        if mapping is None:
            continue
        chain_key, client_full = mapping
        sheets_parsed += 1

        for abs_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None or v == "" for v in row):
                continue
            if len(row) < 18:
                continue

            nr_comanda = _as_str(row[1])
            product_code = _as_str(row[4])
            product_name = _as_str(row[5])
            quantity = _as_decimal(row[8])
            amount = _as_decimal(row[11])
            data_liv_raw = _as_str(row[12])
            data_liv = data_liv_raw[:10] if data_liv_raw else None
            remaining_qty = _as_decimal(row[13])
            status_raw = (str(row[14]).strip().upper() if row[14] is not None else "")
            status = status_raw.replace("NEFACTURARE", "NEFACTURAT")
            category_code = _as_str(row[15])
            ind_val = row[16]
            ship_to = _as_str(row[17])

            if amount is None:
                continue
            if not product_code and not product_name:
                continue
            if not status:
                status = "NEFACTURAT"
            if status not in ("NELIVRAT", "NEFACTURAT"):
                errors.append(
                    f"Sheet '{ws.title}' linia {abs_idx}: status necunoscut "
                    f"'{status_raw}' — skipat"
                )
                continue

            # NELIVRAT: skip dacă remaining ≤ 0 (deja livrat, doar așteaptă
            # facturare). NEFACTURAT: include mereu (remaining poate fi 0).
            if status == "NELIVRAT":
                if remaining_qty is None or remaining_qty <= 0:
                    continue

            # remaining_amount proporțional cu cant_rest / cant (doar NELIVRAT;
            # NEFACTURAT primește suma completă ca remaining).
            remaining_amount: Decimal | None
            if status == "NELIVRAT" and quantity and quantity > 0 and remaining_qty is not None:
                remaining_amount = (remaining_qty / quantity) * amount
            else:
                remaining_amount = amount

            has_ind = bool(
                ind_val is not None and str(ind_val).strip()
                and str(ind_val).strip().lower() not in ("nan", "none")
            )
            ind_str = str(ind_val).strip() if has_ind else None

            # Cheia client pentru match cu alias-uri: "CLIENT | SHIP-TO" (same
            # pattern ca la raw_sales din sales importer).
            client_key = f"{client_full} | {ship_to}" if ship_to else client_full

            effective_code = product_code or product_name

            rows_out.append({
                "source": "adp",
                "chain": chain_key,
                "client": client_key,
                "client_code": None,
                "ship_to": ship_to,
                "nr_comanda": nr_comanda,
                "product_code": effective_code,
                "product_name": product_name,
                "category_code": category_code,
                "status": status,
                "amount": amount,
                "quantity": quantity,
                "remaining_amount": remaining_amount,
                "remaining_quantity": remaining_qty,
                "data_livrare": data_liv,
                "ind": ind_str,
                "has_ind": has_ind,
            })

    if sheets_parsed == 0:
        return [], [
            "Nu s-a găsit niciun sheet KA valid (așteptat: Dedeman, Altex, "
            "Leroy Merlin, Hornbach)."
        ]

    return rows_out, errors
